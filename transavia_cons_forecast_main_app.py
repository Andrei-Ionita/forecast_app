# -*- coding: utf-8 -*-
"""Transavia_Cons_Forecast_main_app.ipynb

Automatically generated by Colaboratory.

Original file is located at
    https://colab.research.google.com/drive/1FhjjUiJlKlpDuJkz5LVX2rE5UR4fORtE

# Setting up the libraries
"""

# Installing xlswriter
!pip install xlsxwriter
# !pip install pystan~=2.14
!pip install greykite
!pip install pyyaml==5.4.1
!pip install anvil-uplink

# Importing the libtraries and tools
import numpy as np
import matplotlib.pyplot as plt
from matplotlib import pyplot
import plotly
import pandas as pd
from sklearn.model_selection import train_test_split
import openpyxl
import xlsxwriter
import xgboost as xgb
import joblib
from pathlib import Path
from sklearn.metrics import mean_squared_error
from sklearn.metrics import mean_absolute_percentage_error
import os
from collections import OrderedDict
import seaborn as sns
from sklearn.metrics import mean_absolute_error
import statistics
import math
from lightgbm import LGBMRegressor
from scipy.stats import randint as sp_randint
from scipy.stats import uniform as sp_uniform
from sklearn.ensemble import RandomForestRegressor
import tensorflow as tf
from greykite.framework.templates.autogen.forecast_config import *
from greykite.framework.templates.forecaster import Forecaster
from greykite.framework.templates.model_templates import ModelTemplateEnum
from greykite.common.features.timeseries_features import *
from greykite.common.evaluation import EvaluationMetricEnum
from greykite.framework.utils.result_summary import summarize_grid_search_results
from greykite.algo.forecast.silverkite.constants.silverkite_holiday import SilverkiteHoliday
from plotly.offline import iplot
import anvil.server
import anvil.media
from google.colab import files
from anvil.google.drive import app_files
import zipfile
import io

"""#Anvil Setup"""

# Connecting to the Anvil app through Uplink
anvil.server.connect("server_LJUDITVW5TYYP5CRMPA4B7A2-ZZFLRJKYDYNBKZXL")

# Creating the dictionary of holidays
New_year_and_day_after = pd.DataFrame({"holiday": "Anul Nou & A doua zi",
                            "ds": pd.to_datetime(["2017-01-01", "2017-01-02", "2016-01-01", "2016-01-02", "2015-01-01", "2015-01-02", "2014-01-01", "2014-01-02", "2019-01-01",
                                                  "2019-01-02", "2018-01-01", "2018-01-02", "2020-01-01", "2020-01-02", "2021-01-01", "2021-01-02",
                                                  "2022-01-01", "2022-01-02", "2023-01-01", "2023-01-02"]),
                            "lower_window": -1,
                            "upper_window": 1})

National_holiday = pd.DataFrame({"holiday": "Ziua Nationala",
                                 "ds": pd.to_datetime(["2016-12-01", "2015-12-01", "2014-12-01", "2018-12-01", "2019-12-01", "2020-12-01", "2021-12-01", "2022-12-01", "2023-12-01"]),
                                 "lower_window": 0,
                                 "upper_window": 1})
Ziua_Principatelor = pd.DataFrame({"holiday": "Ziua Principatelor",
                                 "ds": pd.to_datetime(["2017-01-24", "2016-01-24", "2018-01-24", "2019-01-24", "2020-01-24", "2021-01-24", "2022-01-24", "2023-01-24"]),
                                 "lower_window": 0,
                                 "upper_window": 1})
Christmas = pd.DataFrame({"holiday": "Craciunul",
                          "ds": pd.to_datetime(["2017-12-25", "2017-12-26", "2016-12-25", "2016-12-26", "2015-12-25", "2015-12-26", "2014-12-25", "2014-12-26", "2018-12-25", "2018-12-26", "2019-12-25", "2019-12-26", "2020-12-25", "2020-12-26", "2021-12-25", "2021-12-26",
                                                "2022-12-25", "2022-12-26", "2023-12-25", "2023-12-26"]),
                          "lower_window": -1,
                          "upper_window": 1})
St_Andrew = pd.DataFrame({"holiday": "Sfantul Andrei",
                          "ds": pd.to_datetime(["2017-11-30", "2016-11-30", "2015-11-30", "2014-11-30", "2018-11-30", "2019-11-30", "2020-11-30", "2021-11-30", "2022-11-30",
                                                "2023-11-30"]),
                          "lower_window": -1,
                          "upper_window": 0})
Adormirea_Maicii_Domnului = pd.DataFrame({"holiday": "Adormirea Maicii Domnului",
                                          "ds": pd.to_datetime(["2017-08-15", "2016-08-15", "2015-08-15", "2014-08-15", "2018-08-15", "2019-08-15", "2020-08-15", "2021-08-15","2022-08-15", "2023-08-15"])})
Rusalii = pd.DataFrame({"holiday": "Rusalii",
                        "ds": pd.to_datetime(["2017-06-04", "2017-06-05", "2016-06-19", "2016-06-20", "2015-05-31", "2015-06-01", "2014-06-08", "2014-06-09", "2018-05-27", "2018-05-28", "2019-06-16", "2019-06-17", "2020-06-07", "2020-06-08", "2021-06-20", "2021-06-21",
                                              "2022-06-12", "2022-06-13", "2023-06-04", "2023-06-05"])})
Ziua_Copilului = pd.DataFrame({"holiday": "Ziua Copilului",
                              "ds": pd.to_datetime(["2017-06-01", "2018-06-01", "2019-06-01", "2020-06-01", "2021-06-01", "2022-06-01", "2023-06-01"])})
Ziua_Muncii = pd.DataFrame({"holiday": "Ziua Muncii",
                            "ds": pd.to_datetime(["2017-05-01", "2016-05-01", "2015-05-01", "2014-05-01", "2018-05-01", "2019-05-01", "2020-05-01", "2021-05-01", "2022-05-01", "2023-05-01"])})
Pastele = pd.DataFrame({"holiday": "Pastele",
                        "ds": pd.to_datetime(["2017-04-16", "2017-04-17", "2016-05-01", "2016-05-02", "2015-04-12", "2015-04-13", "2014-04-20", "2014-04-21", "2018-04-08", "2018-04-09", "2019-04-28", "2019-04-29", "2020-04-19", "2020-04-20", "2021-05-02", "2021-05-03",
                                              "2022-04-24", "2022-04-25", "2023-04-16", "2023-04-17"]),
                        "lower_window": -1,
                        "upper_window": 1})
Vinerea_Mare = pd.DataFrame({"holiday": "Vinerea Mare",
                             "ds": pd.to_datetime(["2020-04-17", "2019-04-26", "2018-04-06", "2021-04-30", "2022-04-30", "2023-04-30"])})
Ziua_Unirii = pd.DataFrame({"holiday": "Ziua Unirii",
                            "ds": pd.to_datetime(["2015-01-24", "2020-01-24", "2019-01-24", "2021-01-24", "2022-01-24", "2023-01-24"])})
Public_Holiday = pd.DataFrame({"holiday": "Public Holiday",
                              "ds": pd.to_datetime(["2019-04-30"])})
holidays = pd.concat((New_year_and_day_after, National_holiday, Christmas, St_Andrew, Ziua_Principatelor, Adormirea_Maicii_Domnului, Rusalii, Ziua_Copilului, Ziua_Muncii,
                      Pastele, Vinerea_Mare, Ziua_Unirii, Public_Holiday))

"""# Forecasting Brasov area"""

@anvil.server.callable
def predicting_exporting_consumption_Brasov(dataset):
  # Importing the dataset
  with anvil.media.TempFile(dataset) as f:
    dataset_forecast = pd.read_excel(f, sheet_name="Forecast_data")
    PODs = dataset_forecast.POD.unique()
    datasets_forecast = {elem : pd.DataFrame for elem in PODs}
    for POD in PODs:
      datasets_forecast[POD] = dataset_forecast[:][dataset_forecast.POD == POD]
      datasets_forecast[POD]["WeekDay"] = datasets_forecast[POD].Data.dt.weekday
      datasets_forecast[POD]["Month"] = datasets_forecast[POD].Data.dt.month
      datasets_forecast[POD]["Holiday"] = 0
      for holiday in datasets_forecast[POD]["Data"].unique():
          if holiday in holidays.ds.values:
              datasets_forecast[POD]["Holiday"][datasets_forecast[POD]["Data"] == holiday] = 1
      if len(datasets_forecast[POD]["Flow_Chicks"].value_counts()) > 0 and len(datasets_forecast[POD]["PVPP"].value_counts()) > 0:
        datasets_forecast[POD] = datasets_forecast[POD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Flow_Chicks", "PVPP"]]
      elif len(datasets_forecast[POD]["Flow_Chicks"].value_counts()) > 0:
        datasets_forecast[POD] = datasets_forecast[POD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Flow_Chicks"]]
      elif len(datasets_forecast[POD]["PVPP"].value_counts()) > 0:
        datasets_forecast[POD] = datasets_forecast[POD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "PVPP"]]
      else:
        datasets_forecast[POD] = datasets_forecast[POD][["Month", "WeekDay","Holiday", "Interval", "Temperatura"]]
        # datasets_forecast[POD].replace([np.inf, -np.inf], np.nan)
        # datasets_forecast[POD].dropna(inplace = True)
  # Predicting noPVPP
  predictions = {}
  for POD in datasets_forecast.keys():
    if "PVPP" in datasets_forecast[POD].columns:
      xgb_loaded = joblib.load("/content/drive/MyDrive/Colab Notebooks/Transavia/Consumption/Brasov_Models/rs_xgb_{}.pkl".format(POD))
      print("Predicting for {}".format(POD))
      # print(datasets_forecast[POD])
      xgb_preds = xgb_loaded.predict(datasets_forecast[POD].drop("PVPP", axis=1).values)
      predictions[POD] = xgb_preds
      predictions["Data"] = dataset_forecast["Data"]
      predictions["Interval"] = dataset_forecast["Interval"]
    else:
      xgb_loaded = joblib.load("/content/drive/MyDrive/Colab Notebooks/Transavia/Consumption/Brasov_Models/rs_xgb_{}.pkl".format(POD))
      print("Predicting for {}".format(POD))
      # print(datasets_forecast[POD])
      xgb_preds = xgb_loaded.predict(datasets_forecast[POD].values)
      predictions[POD] = xgb_preds
      predictions["Data"] = dataset_forecast["Data"]
      predictions["Interval"] = dataset_forecast["Interval"]
  # Predicting with PVPP
  predictions_PVPP = {}
  for POD in datasets_forecast.keys():
    if os.path.isfile("/content/drive/MyDrive/Colab Notebooks/Transavia/Consumption/Brasov_Models/rs_xgb_{}_PVPP.pkl".format(POD)):
      xgb_loaded = joblib.load("/content/drive/MyDrive/Colab Notebooks/Transavia/Consumption/Brasov_Models/rs_xgb_{}_PVPP.pkl".format(POD))
      print("Predicting for {}".format(POD))
      # print(datasets_forecast[POD])
      xgb_preds = xgb_loaded.predict(datasets_forecast[POD].values)
      predictions_PVPP[POD] = xgb_preds
      predictions_PVPP["Data"] = dataset_forecast["Data"]
      predictions_PVPP["Interval"] = dataset_forecast["Interval"]
  # Exporting Results to Excel
  workbook = xlsxwriter.Workbook("/content/drive/MyDrive/Colab Notebooks/Transavia/Consumption/Results/XGB/Results_IBDs_daily_Brasov.xlsx")
  worksheet = workbook.add_worksheet("Prediction_Consumption")

  worksheet.write(0,0,"Data")
  worksheet.write(0,1,"Interval")
  worksheet.write(0,2,"Prediction")
  worksheet.write(0,3,"POD")
  worksheet.write(0,4,"Lookup")
  worksheet.write(0,5,"Prediction_PVPP")
  date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
  row = 1
  col = 0
  for POD in datasets_forecast.keys():
    if POD in predictions_PVPP.keys():
      for value in predictions[POD]:
        worksheet.write(row, col+2, abs(value))
        worksheet.write(row, col+3, str(POD))
        row +=1
      row = row - len(predictions[POD])
      for value in predictions_PVPP[POD]:
        worksheet.write(row, col+5, abs(value))
        row +=1
    else:
      for value in predictions[POD]:
        worksheet.write(row, col+2, abs(value))
        worksheet.write(row, col+3, str(POD))
        row +=1
  row = row - (len(predictions[POD]) * len(datasets_forecast.keys()))
  for data in predictions["Data"]:
    worksheet.write(row, col, datetime.date(data),date_format)
    worksheet.write_formula(row, col+4, "=A"+ str(row+1) + "&" + "B"+ str(row+1) + "&" +  "D" + str(row+1))
    row +=1
  row = row - len(predictions["Data"])
  for interval in predictions["Interval"]:
    worksheet.write(row, col+1, interval)
    row +=1
    # row = 1
    # for value in y_test:
    #     worksheet.write(row, col + 1, value)
    #     row +=1
  workbook.close()

"""#Forecast Santimbru Area"""

@anvil.server.callable
def predicting_exporting_consumption_Santimbru(dataset):
  # Importing the dataset
  with anvil.media.TempFile(dataset) as f:
    dataset_forecast = pd.read_excel(f, sheet_name="Forecast_data")
  IBDs = dataset_forecast.IBD.unique()
  IBDs_PVPP = ["Abator", "Abator_Bocsa", "Ciugud", "F5", "Ferma_Bocsa", "FNC", "F4"]
  datasets_forecast = {elem : pd.DataFrame for elem in IBDs}
  for IBD in IBDs:
    datasets_forecast[IBD] = dataset_forecast[:][dataset_forecast.IBD == IBD]
    datasets_forecast[IBD]["WeekDay"] = datasets_forecast[IBD].Data.dt.weekday
    datasets_forecast[IBD]["Month"] = datasets_forecast[IBD].Data.dt.month
    datasets_forecast[IBD]["Holiday"] = 0
    for holiday in datasets_forecast[IBD]["Data"].unique():
        if holiday in holidays.ds.values:
            datasets_forecast[IBD]["Holiday"][datasets_forecast[IBD]["Data"] == holiday] = 1
    if len(datasets_forecast[IBD]["Flow_Chicks"].value_counts()) > 0 and len(datasets_forecast[IBD]["Radiatie"].value_counts()) > 0:
      ## Restructuring the dataset
      datasets_forecast[IBD] = datasets_forecast[IBD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Flow_Chicks", "Radiatie"]]
    elif len(datasets_forecast[IBD]["Flow_Chicks"].value_counts()) > 0:
      datasets_forecast[IBD] = datasets_forecast[IBD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Flow_Chicks"]]
    elif len(datasets_forecast[IBD]["Radiatie"].value_counts()) > 0:
      datasets_forecast[IBD] = datasets_forecast[IBD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Radiatie"]]
    else:
      datasets_forecast[IBD] = datasets_forecast[IBD][["Month", "WeekDay","Holiday", "Interval", "Temperatura"]]
    # datasets_forecast[IBD].replace([np.inf, -np.inf], np.nan)
    # datasets_forecast[IBD].dropna(inplace = True)
    # Check if the cons place has PVPP and add the column, if it does
    if IBD in IBDs_PVPP:
      datasets_forecast[IBD]["PVPP"] = 1
  # Predicting
  predictions = {}
  for IBD in datasets_forecast.keys():
    if IBD in IBDs_PVPP:
      xgb_loaded = joblib.load("/content/drive/MyDrive/Colab Notebooks/Transavia/Consumption/rs_xgb_{}.pkl".format(IBD))
      print("Predicting for {}".format(IBD))
      # print(datasets_forecast[IBD])
      xgb_preds = xgb_loaded.predict(datasets_forecast[IBD].drop("PVPP", axis=1).values)
      predictions[IBD] = xgb_preds
      predictions["Data"] = dataset_forecast["Data"]
      predictions["Interval"] = dataset_forecast["Interval"]
    else:
      xgb_loaded = joblib.load("/content/drive/MyDrive/Colab Notebooks/Transavia/Consumption/rs_xgb_{}.pkl".format(IBD))
      print("Predicting for {}".format(IBD))
      # print(datasets_forecast[IBD])
      xgb_preds = xgb_loaded.predict(datasets_forecast[IBD].values)
      predictions[IBD] = xgb_preds
      predictions["Data"] = dataset_forecast["Data"]
      predictions["Interval"] = dataset_forecast["Interval"]
  # Predicting with PVPP models
  predictions_PVPP = {}
  for IBD in datasets_forecast.keys():
    if IBD in IBDs_PVPP:
      if os.path.isfile("/content/drive/MyDrive/Colab Notebooks/Transavia/Consumption/rs_xgb_{}_PVPP.pkl".format(IBD)):
        xgb_loaded = joblib.load("/content/drive/MyDrive/Colab Notebooks/Transavia/Consumption/rs_xgb_{}_PVPP.pkl".format(IBD))
        print("Predicting for {}".format(IBD))
        # print(datasets_forecast[IBD])
        xgb_preds = xgb_loaded.predict(datasets_forecast[IBD].drop("Radiatie", axis=1).values)
        predictions_PVPP[IBD] = xgb_preds
        predictions_PVPP["Data"] = dataset_forecast["Data"]
        predictions_PVPP["Interval"] = dataset_forecast["Interval"]
  # Exporting Results to Excel
  workbook = xlsxwriter.Workbook("/content/drive/MyDrive/Colab Notebooks/Transavia/Consumption/Results/XGB/Results_IBDs_daily.xlsx")
  worksheet = workbook.add_worksheet("Prediction_Consumption")

  worksheet.write(0,0,"Data")
  worksheet.write(0,1,"Interval")
  worksheet.write(0,2,"Prediction")
  worksheet.write(0,3,"IBD")
  worksheet.write(0,4,"Lookup")
  worksheet.write(0,5,"PVPP_Model_Prediction")
  date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
  row = 1
  col = 0
  for IBD in datasets_forecast.keys():
    if IBD in predictions_PVPP.keys():
      for value in predictions[IBD]:
        worksheet.write(row, col+2, value)
        worksheet.write(row, col+3, IBD)
        row +=1
      row = row-len(predictions[IBD])
      for value in predictions_PVPP[IBD]:
        worksheet.write(row, col+5, value)
        row +=1
    else:
      for value in predictions[IBD]:
        worksheet.write(row, col+2, value)
        worksheet.write(row, col+3, IBD)
        row +=1
  row = row - len(predictions[IBD])*len(datasets_forecast.keys())
  for data in predictions["Data"]:
    worksheet.write(row, col, datetime.date(data),date_format)
    worksheet.write_formula(row, col+4, "=A"+ str(row+1)+ "&" + "B"+ str(row+1)+ "&" + "D" + str(row+1))
    row +=1
  row = row - len(predictions["Data"])
  for interval in predictions["Interval"]:
    worksheet.write(row, col+1, interval)
    row +=1
    # row = 1
    # for value in y_test:
    #     worksheet.write(row, col + 1, value)
    #     row +=1
  workbook.close()

"""# Solina Production Forecast"""

@anvil.server.callable
# Predicting on Forecast Dataset and exporting the results
def predicting_exporting(dataset):
  with anvil.media.TempFile(dataset) as f:
    dataset_forecast = pd.read_excel(f, sheet_name="Forecast_Dataset")
  xgb_loaded = joblib.load("/content/drive/MyDrive/Colab Notebooks/Solina/rs_xgb_Solina_prod.pkl")
  dataset_forecast["Month"] = dataset_forecast.Data.dt.month

  dataset_forecast = dataset_forecast.drop("Data", axis=1)

  preds = xgb_loaded.predict(dataset_forecast.values)
  #Exporting Results to Excel
  workbook = xlsxwriter.Workbook("/content/drive/MyDrive/Colab Notebooks/Solina/Results_Production_xgb.xlsx")
  worksheet = workbook.add_worksheet("Production_Predictions")

  row = 1
  col = 0
  worksheet.write(0,0,"Prediction")
  # worksheet.write(0,1,"Real")

  for value in preds:
      worksheet.write(row, col, value)
      row +=1
  # row = 1
  # for value in y_test:
  #     worksheet.write(row, col + 1, value)
  #     row +=1

  workbook.close()

"""# Transavia Production Forecast"""

@anvil.server.callable
# Predicting on Forecast Dataset and exporting the results
def predicting_exporting_Transavia(dataset):
  with anvil.media.TempFile(dataset) as f:
    dataset_forecast = pd.read_excel(f)
  CEFs = dataset_forecast.Centrala.unique()
  datasets_forecast = {elem : pd.DataFrame for elem in CEFs}
  for CEF in CEFs:
    print("Predicting for {}".format(CEF))
    xgb_loaded = joblib.load("/content/drive/MyDrive/Colab Notebooks/Transavia/Production/Models/rs_xgb_{}.pkl".format(CEF))
    dataset_forecast = pd.read_excel("/content/drive/MyDrive/Colab Notebooks/Transavia/Production/Input_production.xlsx")
    dataset_forecast = dataset_forecast[:][dataset_forecast.Centrala == CEF]
    dataset_forecast["Month"] = dataset_forecast.Data.dt.month
    if CEF in ["F24"]:
      df_forecast = dataset_forecast.drop(["Data", "Nori", "Centrala"], axis=1)
    else:
      df_forecast = dataset_forecast.drop(["Data", "Centrala"], axis=1)
    preds = xgb_loaded.predict(df_forecast.values)
    # Exporting Results to Excel
    workbook = xlsxwriter.Workbook("/content/drive/MyDrive/Colab Notebooks/Transavia/Production/Results/Results_daily_{}.xlsx".format(CEF))
    worksheet = workbook.add_worksheet("Prediction_Production")

    worksheet.write(0,0,"Data")
    worksheet.write(0,1,"Interval")
    worksheet.write(0,2,"Production")
    date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
    row = 1
    col = 0

    for value in preds:
      worksheet.write(row, col+2, value)
      row +=1
    row = row - len(preds)
    for data in dataset_forecast["Data"]:
      worksheet.write(row, col, data, date_format)
      row +=1
    row = row - len(dataset_forecast["Data"])
    for value in dataset_forecast["Interval"]:
      worksheet.write(row, col+1, value)
      row +=1
    workbook.close()
def zip_files(folder_path, zip_name):
    zip_path = os.path.join(folder_path, zip_name)
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for root, _, files in os.walk(folder_path):
            for file in files:
                if file != zip_name:  # Avoid zipping the zip file itself
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, folder_path)  # Relative path within the zip file
                    zipf.write(file_path, arcname)

# Usage:
folder_path = '/content/drive/MyDrive/Colab Notebooks/Transavia/Production/Results'
zip_name = 'Transavia_Production_Results.zip'
zip_files(folder_path, zip_name)

anvil.server.wait_forever()

while True:pass

!pip freeze > requirements.txt