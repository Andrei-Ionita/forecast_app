import streamlit as st
import pandas as pd
import numpy as np
import json
import requests
import xlsxwriter
from datetime import datetime, timedelta
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle, numbers
# import win32com.client as win32
# import pythoncom
# from pywintypes import com_error
import os, os.path
# import win32com.client
import joblib
import base64
from pathlib import Path
import time

# Creating the holidays dataframe
# Creating the dictionary of holidays
New_year_and_day_after = pd.DataFrame({"holiday": "Anul Nou & A doua zi",
                                                        "ds": pd.to_datetime(["2017-01-01", "2017-01-02", "2016-01-01", "2016-01-02", "2015-01-01", "2015-01-02", "2014-01-01", "2014-01-02", "2019-01-01",
                                                                                                    "2019-01-02", "2018-01-01", "2018-01-02", "2020-01-01", "2020-01-02", "2021-01-01", "2021-01-02",
                                                                                                    "2022-01-01", "2022-01-02", "2023-01-01", "2023-01-02","2024-01-01", "2024-01-02"]),
                                                        "lower_window": -1,
                                                        "upper_window": 1}) 

National_holiday = pd.DataFrame({"holiday": "Ziua Nationala",
                                                                 "ds": pd.to_datetime(["2016-12-01", "2015-12-01", "2014-12-01", "2018-12-01", "2019-12-01", "2020-12-01", "2021-12-01", "2022-12-01", "2023-12-01", "2024-12-01"]),
                                                                 "lower_window": 0,
                                                                 "upper_window": 1})
Ziua_Principatelor = pd.DataFrame({"holiday": "Ziua Principatelor",
                                                                 "ds": pd.to_datetime(["2017-01-24", "2016-01-24", "2018-01-24", "2019-01-24", "2020-01-24", "2021-01-24", "2022-01-24", "2023-01-24", "2024-01-24"]),
                                                                 "lower_window": 0,
                                                                 "upper_window": 1})
Christmas = pd.DataFrame({"holiday": "Craciunul",
                                                    "ds": pd.to_datetime(["2017-12-25", "2017-12-26", "2016-12-25", "2016-12-26", "2015-12-25", "2015-12-26", "2014-12-25", "2014-12-26", "2018-12-25", "2018-12-26", "2019-12-25", "2019-12-26", "2020-12-25", "2020-12-26", "2021-12-25", "2021-12-26",
                                                                                                "2022-12-25", "2022-12-26", "2023-12-25", "2023-12-26", "2024-12-25", "2024-12-26"]),
                                                    "lower_window": -1,
                                                    "upper_window": 1})
St_Andrew = pd.DataFrame({"holiday": "Sfantul Andrei",
                                                    "ds": pd.to_datetime(["2017-11-30", "2016-11-30", "2015-11-30", "2014-11-30", "2018-11-30", "2019-11-30", "2020-11-30", "2021-11-30", "2022-11-30",
                                                                                                "2023-11-30", "2024-11-30"]),
                                                    "lower_window": -1,
                                                    "upper_window": 0})
Adormirea_Maicii_Domnului = pd.DataFrame({"holiday": "Adormirea Maicii Domnului",
                                                                                    "ds": pd.to_datetime(["2017-08-15", "2016-08-15", "2015-08-15", "2014-08-15", "2018-08-15", "2019-08-15", "2020-08-15", "2021-08-15","2022-08-15", "2024-08-15"])})
Rusalii = pd.DataFrame({"holiday": "Rusalii",
                                                "ds": pd.to_datetime(["2017-06-04", "2017-06-05", "2016-06-19", "2016-06-20", "2015-05-31", "2015-06-01", "2014-06-08", "2014-06-09", "2018-05-27", "2018-05-28", "2019-06-16", "2019-06-17", "2020-06-07", "2020-06-08", "2021-06-20", "2021-06-21",
                                                                                            "2022-06-12", "2022-06-13", "2023-06-04", "2023-06-05", "2024-06-24"])})
Ziua_Copilului = pd.DataFrame({"holiday": "Ziua Copilului",
                                                            "ds": pd.to_datetime(["2017-06-01", "2018-06-01", "2019-06-01", "2020-06-01", "2021-06-01", "2022-06-01", "2023-06-01", "2024-06-01"])})
Ziua_Muncii = pd.DataFrame({"holiday": "Ziua Muncii",
                                                        "ds": pd.to_datetime(["2017-05-01", "2016-05-01", "2015-05-01", "2014-05-01", "2018-05-01", "2019-05-01", "2020-05-01", "2021-05-01", "2022-05-01", "2023-05-01",
                                                            "2024-05-01"])})
Pastele = pd.DataFrame({"holiday": "Pastele",
                                                "ds": pd.to_datetime(["2017-04-16", "2017-04-17", "2016-05-01", "2016-05-02", "2015-04-12", "2015-04-13", "2014-04-20", "2014-04-21", "2018-04-08", "2018-04-09", "2019-04-28", "2019-04-29", "2020-04-19", "2020-04-20", "2021-05-02", "2021-05-03",
                                                                                            "2022-04-24", "2022-04-25", "2023-04-16", "2023-04-17", "2024-05-06"]),
                                                "lower_window": -1,
                                                "upper_window": 1})
Vinerea_Mare = pd.DataFrame({"holiday": "Vinerea Mare",
                                                         "ds": pd.to_datetime(["2020-04-17", "2019-04-26", "2018-04-06", "2021-04-30", "2022-04-30", "2023-04-30", "2024-05-03"])})
Ziua_Unirii = pd.DataFrame({"holiday": "Ziua Unirii",
                                                        "ds": pd.to_datetime(["2015-01-24", "2020-01-24", "2019-01-24", "2021-01-24", "2022-01-24", "2023-01-24", "2024-01-24"])})
Public_Holiday = pd.DataFrame({"holiday": "Public Holiday",
                                                            "ds": pd.to_datetime(["2019-04-30"])})
holidays = pd.concat((New_year_and_day_after, National_holiday, Christmas, St_Andrew, Ziua_Principatelor, Adormirea_Maicii_Domnului, Rusalii, Ziua_Copilului, Ziua_Muncii,
                                            Pastele, Vinerea_Mare, Ziua_Unirii, Public_Holiday))

# Defining the functions
def fetch_ghi_data_from_api(lat, lon, date, api_key):
    # Fetch data from the API
    api_url = "https://api.openweathermap.org/energy/1.0/solar/data?lat={}&lon={}&date={}&appid={}".format(lat, lon, date, api_key)
    response = requests.get(api_url)
    if response.status_code == 200:
        return response.json()
    else:
        raise Exception(f"Failed to fetch data: Status code {response.status_code}")

def fetch_temps_clouds_data_from_api(lat, lon, api_key, nr_days):
    # Fetch data from the API
    api_url = api_url = "https://pro.openweathermap.org/data/2.5/forecast/hourly?lat={}&lon={}&appid={}&cnt={}&units=metric".format(lat, lon, api_key, nr_days)
    response = requests.get(api_url)
    if response.status_code == 200:
        return response.json()
    else:
        raise Exception(f"Failed to fetch data: Status code {response.status_code}")

def ghi_json_to_excel(ghi_data, output_file_path):
    # Load JSON data
    json_data = ghi_data

    # Flattening 'irradiance' data for 'daily' and 'hourly'
    daily_irradiance = pd.DataFrame(json_data['irradiance']['daily'])
    hourly_irradiance = pd.DataFrame(json_data['irradiance']['hourly'])

    # Further flatten nested data in 'irradiance'
    daily_clear_sky = pd.json_normalize(daily_irradiance['clear_sky']).add_prefix('clear_sky_')
    daily_cloudy_sky = pd.json_normalize(daily_irradiance['cloudy_sky']).add_prefix('cloudy_sky_')
    flattened_daily_irradiance = pd.concat([daily_clear_sky, daily_cloudy_sky], axis=1)

    hourly_irradiance['clear_sky'] = hourly_irradiance['clear_sky'].apply(lambda x: {} if pd.isna(x) else x)
    hourly_irradiance['cloudy_sky'] = hourly_irradiance['cloudy_sky'].apply(lambda x: {} if pd.isna(x) else x)
    hourly_clear_sky = pd.json_normalize(hourly_irradiance['clear_sky']).add_prefix('clear_sky_')
    hourly_cloudy_sky = pd.json_normalize(hourly_irradiance['cloudy_sky']).add_prefix('cloudy_sky_')
    flattened_hourly_irradiance = pd.concat([hourly_irradiance['hour'], hourly_clear_sky, hourly_cloudy_sky], axis=1)

    # Extracting top-level data, excluding the 'irradiance' key
    top_level_data = {k: json_data[k] for k in json_data if k != 'irradiance'}
    top_level_df = pd.DataFrame([top_level_data])

    # Merging the top-level data with the flattened irradiance data
    merged_daily = pd.concat([top_level_df] * len(flattened_daily_irradiance), ignore_index=True)
    merged_daily = pd.concat([merged_daily, flattened_daily_irradiance], axis=1)
    merged_hourly = pd.concat([top_level_df] * len(flattened_hourly_irradiance), ignore_index=True)
    merged_hourly = pd.concat([merged_hourly, flattened_hourly_irradiance], axis=1)

    # Saving to Excel
    with pd.ExcelWriter(output_file_path) as writer:
        merged_daily.to_excel(writer, sheet_name='Daily', index=False)
        merged_hourly.to_excel(writer, sheet_name='Hourly', index=False)

def save_json_to_file(json_data, folder_path, file_name):
    # Create folder if it doesn't exist
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    # Full path for the json file
    file_path = os.path.join(folder_path, file_name)

    # Write the JSON data to the file
    with open(file_path, 'w') as file:
        json.dump(json_data, file)

    print(f"File saved successfully at {file_path}")

def flatten_json_to_excel(json_file_path, output_file_path):
    # Load JSON data
    with open(json_file_path, 'r') as file:
        raw_json_data = json.load(file)

    # Extract the data under the 'list' key
    list_data = raw_json_data['list']

    # Flattening the 'list' data
    flattened_records = []
    for record in list_data:
        # Flattening top-level fields
        flattened_record = pd.json_normalize(record)
        # Appending the flattened record to the list
        flattened_records.append(flattened_record)

    # Concatenate all flattened records into a DataFrame
    flattened_list_data = pd.concat(flattened_records, ignore_index=True)

    # Saving to Excel
    flattened_list_data.to_excel(output_file_path, index=False)

def concatenate_ghi_data(CEF, forecast):
    folder_path = './{}/Production/Input/'.format(CEF)  # Replace with your folder path
    specific_string = 'GHI_'  # Replace with the string you want to check for
    all_files = [f for f in os.listdir(folder_path) if specific_string in f]


    # Lists to hold DataFrames from each sheet
    all_daily_dataframes = []
    all_hourly_dataframes = []

    for filename in all_files:
        file_path = os.path.join(folder_path, filename)
        daily_df = pd.read_excel(file_path, sheet_name='Daily')
        hourly_df = pd.read_excel(file_path, sheet_name='Hourly')

        all_daily_dataframes.append(daily_df)
        all_hourly_dataframes.append(hourly_df)

    # Concatenate DataFrames from each sheet
    concatenated_daily_df = pd.concat(all_daily_dataframes, ignore_index=True)
    concatenated_hourly_df = pd.concat(all_hourly_dataframes, ignore_index=True)

    # Optionally, save the concatenated DataFrames to new Excel files
    # concatenated_daily_df.to_excel('Concatenated_Daily_GHI.xlsx', index=False)
    concatenated_hourly_df.to_excel('./{}/{}/Input/Concatenated_Hourly_GHI.xlsx'.format(CEF, forecast), index=False)

    print("All files concatenated into Concatenated_Daily_GHI.xlsx and Concatenated_Hourly_GHI.xlsx")


#=======================================================================Preprocessing GHI data======================================================================================

# Setup for GHI
lat = 47.2229
lon = 24.7244

api_key = "f62968d774964bad9bee981406d43d8e"
date = datetime.now().strftime('%Y-%m-%d')

# Fetching the data
ghi_data = fetch_ghi_data_from_api(lat, lon, date, api_key)

# Preprocessing the data
output_file_path = './RAAL/Weather_data/GHI_{}.xlsx'.format(date)  # Replace with your desired output file path
ghi_json_to_excel(ghi_data, output_file_path)

# Formatting the date column (Column "C")
def process_and_save_excel(file_path):
    # Load dataframe
    df = pd.read_excel(file_path, engine='openpyxl')
    # Convert 'date' column to datetime type
    df['date'] = pd.to_datetime(df['date'])
    # Create a Pandas Excel writer using XlsxWriter as the engine.
    writer = pd.ExcelWriter(file_path, engine='xlsxwriter') 
    # Convert the dataframe to an XlsxWriter Excel object.
    df.to_excel(writer, sheet_name='Sheet1', index=False)    
    # Get workbook
    workbook  = writer.book
    # Add a format for the date using xlsxwriter
    date_format = workbook.add_format({'num_format': 'dd.mm.yyyy'})
    # Get the worksheet
    worksheet = writer.sheets['Sheet1']
    # Set date format
    worksheet.set_column('C:C', None, date_format)
    # Close the Pandas Excel writer and output the Excel file
    workbook.close()

# Defining the lookup column in the GHI file
def add_lookup_column_GHI():
    # Load the workbook and sheet
    workbook = load_workbook("./RAAL/Production/Input/Concatenated_Hourly_GHI.xlsx")
    sheet = workbook.active

    # Determine the last column with data
    last_col = sheet.max_column
    
    # Create a new column header 'Lookup' in the next column
    lookup_col_letter = get_column_letter(last_col + 1)
    sheet[f"{lookup_col_letter}1"] = 'Lookup'
    
    # Determine the last row with data in column C
    last_row = max((c.row for c in sheet['C'] if c.value is not None))

    # Starting from the second row, insert the formula
    for row in range(2, last_row + 1):
        sheet[f"{lookup_col_letter}{row}"] = f"=C{row}&G{row}"

    # Save the workbook
    workbook.save("./RAAL/Production/Input/Concatenated_Hourly_GHI.xlsx")


def prepare_lookup_column(file_path, date_col_name='Data', interval_col_name='Interval'):
    # Load the Excel file into a DataFrame
    df = pd.read_excel(file_path)
    
    # Ensure the 'Data' column is in datetime format
    df[date_col_name] = pd.to_datetime(df[date_col_name])
    
    # Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
    # Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
    df['Lookup'] = df[date_col_name].dt.strftime('%d.%m.%Y') + df[interval_col_name].astype(str)
    df.to_excel(file_path, index=False)
    return df


#=======================================================================Preprocessing temps_clouds data======================================================================================

# Setup for temps_clouds
api_key = "f62968d774964bad9bee981406d43d8e"
date = datetime.now().strftime('%Y-%m-%d')

# Fetching the data
# temps_clouds_data = fetch_temps_clouds_data_from_api(lat,lon,api_key,nr_days)

# Preprocessing the data
# json_file_path = temps_clouds_data  # Replace with your JSON file path
# output_file_path = './RAAL/Weather_data/temps_clouds_{}.xlsx'.format(date)  # Replace with your desired output file path

# First, we need to do some preprocessing of the weather and GHI file
def creating_weather_date_hour_columns(CEF, forecast):
    # Load the data
    weather_df = pd.read_excel('./{}/{}/Input/weather.xlsx'.format(CEF, forecast))

    # Ensure the column 'E' is in datetime format
    weather_df['dt_txt'] = pd.to_datetime(weather_df['dt_txt'])

    # Insert new columns for Date and Hour
    weather_df.insert(loc=weather_df.columns.get_loc('dt_txt') + 1, column='Data', value=weather_df['dt_txt'].dt.date)
    weather_df.insert(loc=weather_df.columns.get_loc('Data') + 1, column='Interval', value=weather_df['dt_txt'].dt.hour)

    # Save the modified DataFrame back to Excel
    weather_df.to_excel('./{}/{}/Input/weather.xlsx'.format(CEF, forecast), index=False)

def change_date_format_weather_wb(CEF, forecast):
    # Check if the named style already exists
    weather_wb = openpyxl.open('./{}/Production/Input/weather.xlsx'.format(CEF))
    style_name = 'date_style'
    if style_name not in weather_wb.named_styles:
        # Create a new NamedStyle and add it to the workbook
        date_style = NamedStyle(name=style_name, number_format='DD.MM.YYYY')
        weather_wb.add_named_style(date_style)
    else:
        # If the style already exists, there's no need to add it again
        print(f"Style '{style_name}' already exists in the workbook.")
    # Now you can safely assign the style to cells without causing an error
    for row in range(1, weather_wb.active.max_row + 1):
        weather_wb.active.cell(row=row, column=6).style = date_style
    weather_wb.save('./{}/{}/Input/weather.xlsx'.format(CEF, forecast))

# Defining the lookup column in the weather file
def add_lookup_column_weather(CEF, forecast):
    # Load the workbook and sheet
    workbook = load_workbook('./{}/Production/Input/weather.xlsx'.format(CEF))
    sheet = workbook.active

    # Determine the last column with data
    last_col = sheet.max_column
    
    # Create a new column header 'Lookup' in the next column
    lookup_col_letter = get_column_letter(last_col + 1)
    sheet[f"{lookup_col_letter}1"] = 'Lookup'
    
    # Determine the last row with data in column F
    last_row = max((c.row for c in sheet['F'] if c.value is not None))

    # Starting from the second row, insert the formula
    for row in range(2, last_row + 1):
        sheet[f"{lookup_col_letter}{row}"] = f"=F{row}&G{row}"

    # Save the workbook
    workbook.save('./{}/{}/Input/weather.xlsx'.format(CEF, forecast))
    weather_df = pd.read_excel('./{}/{}/Input/weather.xlsx'.format(CEF, forecast))
    weather_df["Lookup_python"] = weather_df["Data"].astype(str) + weather_df["Interval"].astype(str)
    weather_df.to_excel('./{}/{}/Input/weather.xlsx'.format(CEF, forecast))

def prepare_lookup_column(file_path, date_col_name='Data', interval_col_name='Interval'):
    # Load the Excel file into a DataFrame
    df = pd.read_excel(file_path)

    # Ensure the 'Data' column is in datetime format
    df[date_col_name] = pd.to_datetime(df[date_col_name])
    
    # Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
    # Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
    df['Lookup'] = df[date_col_name].dt.strftime('%d.%m.%Y') + df[interval_col_name].astype(str)
    df.to_excel(file_path, index=False)
    return df
#=======================================================================Creating Input file======================================================================================

# Creating the Input file
# workbook = xlsxwriter.Workbook("./RAAL/Production/Input.xlsx")
# worksheet = workbook.add_worksheet("forecast_dataset")
# date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
# row = 1
# col = 0
# worksheet.write(0,0,"Data")
# worksheet.write(0,1,"Interval")
# worksheet.write(0,2,"Radiatie")
# worksheet.write(0,3,"Temperatura")
# worksheet.write(0,4,"Nori")

# # for value in preds:
# #     worksheet.write(row, col + 2, value)
# #     row +=1
# # row = 1
# # for Data, Interval in zip(dataset.Data, dataset.Interval):
# #     worksheet.write(row, col + 0, Data, date_format)
# #     worksheet.write(row, col + 1, Interval)
# #     row +=1

# workbook.close()

# def run_vba_macro():
#     if os.path.exists(file_path):
#             pythoncom.CoInitialize()
#             xl=win32com.client.Dispatch("Excel.Application")
#             xl.Workbooks.Open(os.path.abspath(file_path), ReadOnly=1)
#             print("Workbook Opened")
#             try:
#                 xl.Application.Run("Data_Eng.Building_Input_file")
#             except com_error as e:
#                 print(f"COM error encountered: {e}")
#                 # Handle or log the error as needed. Consider continuing if it's a known non-critical error.
#             print("Macro finished!")
#             # xl.Application.Save() # if you want to save then uncomment this line and change delete the ", ReadOnly=1" part from the open function.
#             # xl.Application.Quit() # Comment this out if your excel script closes
#             del xl
#             pythoncom.CoUninitialize()

# Example usage
file_path = "./RAAL/Production/Input.xlsm"  # Update this path to your actual file path

def building_input_file(CEF, forecast):
    # Define file paths
    # base_path = Path("./RAAL/Production").parent  # Adjust this path as necessary
    ghi_file_path = "./{}/Production/Input/Concatenated_Hourly_GHI.xlsx".format(CEF)
    weather_file_path = "./{}/Production/Input/weather.xlsx".format(CEF)
    
    # Load workbooks and sheets
    ghi_wb = openpyxl.load_workbook(ghi_file_path, data_only=True)
    weather_wb = openpyxl.load_workbook(weather_file_path, data_only=True)
    from pathlib import Path
    # Example for loading an existing workbook
    if forecast == "Production":
        workbook_path = Path("./{}/Production/Input_Production_{}.xlsx".format(CEF, CEF))  # Ensure this path is correct
    else:
        workbook_path = Path("./{}/Consumption/Input_Consumption_{}.xlsx".format(CEF, CEF))
    if workbook_path.exists():
        main_wb = openpyxl.load_workbook(workbook_path)
        ws = main_wb.active  # or specify the sheet name directly if needed

        # Check if the named style already exists
        style_name = 'date_style'
        if style_name not in main_wb.named_styles:
            # Create a new NamedStyle and add it to the workbook
            date_style = NamedStyle(name=style_name, number_format='DD.MM.YYYY')
            main_wb.add_named_style(date_style)
        else:
            # If the style already exists, there's no need to add it again
            print(f"Style '{style_name}' already exists in the workbook.")
    else:
        print(f"Workbook not found at {workbook_path}")
    # Copying weather data
    weather_ws = weather_wb.active  # Adjust the sheet if necessary
    weather_data = pd.read_excel(weather_file_path, sheet_name=weather_ws.title)
    dates_intervals = weather_data[['Data', 'Interval']]  # Adjust column names as necessary
    for index, row in dates_intervals.iterrows():
        # Make sure ws is defined
        if ws is not None:
            # Assuming 'Data' is the correct column name; adjust as needed
            for index, row in dates_intervals.iterrows():
                if 'Data' in row:  # Check if 'Data' column exists in row
                    ws.cell(row=index + 2, column=1, value=row['Data'])
                    ws.cell(row=index + 2, column=2, value=row['Interval'])
                else:
                    print(f"Column 'Data' not found in row {index + 2}")
        else:
            print("Worksheet not initialized.")
    # Assuming you have specific logic to match and insert formulas, here's a simplified example
    # For GHI data
    # last_row = ws.max_row
    # for row in range(2, last_row + 1):
    #     # Constructing the INDEX MATCH formula as a string
    #     # Adjust the formula according to your specific Excel file names, sheet names, and column references
    #     index_match_formula = (
    #         f'=INDEX([Concatenated_Hourly_GHI.xlsx]Sheet1!$K:$K, '
    #         f'MATCH(A{row}&B{row}, [Concatenated_Hourly_GHI.xlsx]Sheet1!$N:$N, 0))'
    #     )

    #     # Setting the formula for a specific cell, e.g., in the 'Radiatie' column (assuming column C)
    #     ws.cell(row=row, column=3).value = index_match_formula
    # Load your main data and the data to be matched against
    # main_df = pd.read_excel(workbook_path)
    # st.write(main_df)
    # lookup_df = pd.read_excel(ghi_file_path)
    # # st.write(lookup_df)
    # # # Create a new column in both dataframes that concatenates the values of columns A and B

    # # Create the 'MatchKey' column by concatenating the formatted 'Data' and 'Interval'
    # Adjust 'ColumnA' and 'ColumnB' to match your actual column names
    # main_df['MatchKey'] = main_df["Lookup"]
    # lookup_df['MatchKey'] = lookup_df["Lookup"].astype(str)
    # st.write(lookup_df)
    # # Now, perform the match based on this new 'MatchKey' column and retrieve the desired value from the lookup dataframe
    # # Adjust 'ValueColumn' to the actual name of the column you want to retrieve after matching
    # main_df['Radiatie'] = main_df['MatchKey'].map(lookup_df.set_index('MatchKey')['cloudy_sky_ghi'])

    # # Similar steps for temperature and clouds from weather data, adjust the formula and columns as necessary
    
    # # Save and close workbooks
    main_wb.save("./{}/{}/Input_{}_{}.xlsx".format(CEF, forecast, forecast, CEF))  # Adjust the path as necessary
    # No need to explicitly close in Python, as workbooks are closed when the program ends or when they're no longer referenced

# Adding the Lookup column to the Input file
def add_lookup_column_Input():

    # # Let's create the Lookup column in the main workbook
    # last_row = ws.max_row
    # ws.cell(1,6).value = "Lookup"
    # for row in range(2, last_row + 1):
    #     # Concatenate the values of columns A and B for each row
    #     ws.cell(row=row, column=6).value = f'=A{row}&B{row}'
    # Load the workbook and sheet
    workbook = load_workbook('./RAAL/Production/Input.xlsx')
    sheet = workbook.active

    # Determine the last column with data
    # last_col = 5
    
    # Create a new column header 'Lookup' in the next column
    # lookup_col_letter = get_column_letter(last_col + 1)
    sheet["E1"] = 'Lookup'
    
    # Determine the last row with data in column F
    last_row = max((c.row for c in sheet['A'] if c.value is not None))

    # Starting from the second row, insert the formula
    for row in range(2, last_row + 1):
        sheet[f"E{row}"] = f"=A{row}&B{row}"

    # Save the workbook
    workbook.save('./RAAL/Production/Input.xlsx')

def add_lookup_column_input_xlsxwriter():
    # Reading the weather file with openpyxl
    weather_df = pd.read_excel("./RAAL/Production/Input/weather.xlsx")
    st.write(weather_df)
    wb = xlsxwriter.Workbook("./RAAL/Production/Input_xlsxwriter.xlsx")
    ws = wb.add_worksheet("forecast_dataset")
    # Adding the columns
    ws.write(0,0,"Data")
    ws.write(0,1,"Interval")
    ws.write(0,2,"Radiatie")
    ws.write(0,3,"Temperatura")
    ws.write(0,4,"Nori")
    ws.write(0,5,"Lookup")
    date_format = wb.add_format({'num_format':'dd.mm.yyyy'})
    # Writing the dates
    for row, data in enumerate(weather_df["Data"], start=1):
        # Convert the string to datetime object
        # This assumes your date format in the DataFrame is consistent and recognized by strptime
        date_obj = datetime.strptime(str(data), "%Y-%m-%d %H:%M:%S")
        ws.write_datetime(row, 0, date_obj, date_format)
    # Writing the dates intervals
    for row, interval in enumerate(weather_df["Interval"], start=1):
        ws.write(row,1,interval)
    # row=1
    # col=0
    # # Creating the Lookup column
    # for data in weather_df["Data"].values:
    #     ws.write_formula(row, col+5, "=A"+ str(row+1) + "&" + "B"+ str(row+1))
    #     row +=1
    row = 1  # Starting row (after headers)
    for index, data in weather_df.iterrows():
        # Write the original data in columns A and B
        ws.write(row, 0, data['Data'], date_format)
        ws.write(row, 1, data['Interval'])

        # Creating the Lookup formula in column F (index 5)
        formula = f'=A{row+1}&B{row+1}'
        ws.write_formula(row, 5, formula)
        
        # Also, write the expected result of the formula as a string in another column (for example, column G)
        # This replicates the formula's action in Python and writes it as a string
        concatenated_result = str(data['Data']) + str(data['Interval'])
        ws.write(row, 6, concatenated_result)
        
        row += 1
    wb.close()

# Lookuping the GHI values
# def lookup_ghi_values():
#     main_df = pd.read_excel("./RAAL/Production/Input.xlsx")
#     st.write(main_df)
#     main_df["Lookup"] = main_df["Lookup"].astype(str)
#     lookup_df = pd.read_excel("./RAAL/Production/Input/Concatenated_Hourly_GHI.xlsx")
#     lookup_df["Lookup"] = lookup_df["Lookup"].astype(str)
#     st.write(lookup_df)
#     # Create a dictionary from lookup_df for efficient lookup
#     lookup_dict = lookup_df.set_index("Lookup")["cloudy_sky_ghi"].to_dict()
#     # Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
#     main_df['Radiatie'] = main_df['Lookup'].map(lookup_dict)
#     # Check the result
#     print(main_df[['Lookup', 'Radiatie']])
#     # Save the updated DataFrame to an Excel file
#     main_df.to_excel('./RAAL/Production/Input.xlsx', index=False)

# Lookuping the temperatures and clouds
def lookup_weather_values():
    main_df = pd.read_excel("./RAAL/Production/Input.xlsx")
    main_df["Lookup"] = main_df["Lookup"].astype(str)
    lookup_df = pd.read_excel("./RAAL/Production/Input/weather.xlsx")
    lookup_df["Lookup"] = lookup_df["Lookup"].astype(str)
    # Temperatures values
    # Create a dictionary from lookup_df for efficient lookup
    lookup_dict = lookup_df.set_index("Lookup")["main.temp"].to_dict()
    # Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
    main_df['Temperatura'] = main_df['Lookup'].map(lookup_dict)
    # Check the result
    # print(main_df[['Lookup', 'Radiatie']])
    # Save the updated DataFrame to an Excel file
    # Clouds values
    # Create a dictionary from lookup_df for efficient lookup
    lookup_dict = lookup_df.set_index("Lookup")["clouds.all"].to_dict()
    # Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
    main_df['Nori'] = main_df['Lookup'].map(lookup_dict)
    main_df.to_excel('./RAAL/Production/Input.xlsx', index=False)

def lookup_ghi_values(CEF, input_df, ghi_df, forecast):
    # Create a dictionary from the GHI DataFrame for efficient lookup
    ghi_dict = ghi_df.set_index('Lookup')['cloudy_sky_ghi'].to_dict()
    
    # Map the 'Lookup' values from the input DataFrame to get the 'cloudy_sky_ghi' values
    input_df['Radiatie'] = input_df['Lookup'].map(ghi_dict)
    input_df.to_excel("./{}/{}/Input_{}_{}.xlsx".format(CEF, forecast, forecast, CEF), index=False)
    return input_df

def lookup_weather_values(CEF, input_df, weather_df, forecast):
    # Create a dictionary from the main.temp DataFrame for efficient lookup
    weather_dict = weather_df.set_index('Lookup')['main.temp'].to_dict()
    
    # Map the 'Lookup' values from the input DataFrame to get the 'temperatures' values
    input_df['Temperatura'] = input_df['Lookup'].map(weather_dict)
    # Create a dictionary from the main.temp DataFrame for efficient lookup
    weather_dict = weather_df.set_index('Lookup')['clouds.all'].to_dict()
    
    # Map the 'Lookup' values from the input DataFrame to get the 'temperatures' values
    input_df['Nori'] = input_df['Lookup'].map(weather_dict)
    input_df.to_excel("./{}/{}/Input_{}_{}.xlsx".format(CEF, forecast, forecast, CEF), index=False)
    return input_df
#===============================================================================Forcasting RAAL Production=================================================================

def predicting_exporting_RAAL(dataset):
    xgb_loaded = joblib.load("./RAAL/Production/rs_xgb_RAAL_prod.pkl")
    dataset.dropna(inplace=True)
    dataset_forecast = dataset.copy()
    dataset_forecast = dataset_forecast[["Data", "Interval", "Radiatie", "Temperatura"]]
    dataset_forecast["Month"] = dataset_forecast.Data.dt.month
    dataset_forecast.dropna(inplace=True)
    dataset_forecast = dataset_forecast.drop("Data", axis=1)
    st.write(dataset_forecast)
    preds = xgb_loaded.predict(dataset_forecast.values)
    #Exporting Results to Excel
    workbook = xlsxwriter.Workbook("./RAAL/Production/Results/Results_Production_xgb_RAAL_wm.xlsx")
    worksheet = workbook.add_worksheet("Production_Predictions")
    date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
    # Define a format for cells with three decimal places
    decimal_format = workbook.add_format({'num_format': '0.000'})
    row = 1
    col = 0
    worksheet.write(0,0,"Data")
    worksheet.write(0,1,"Interval")
    worksheet.write(0,2,"Prediction")
    # Rounded values
    rounded_preds = [round(val, 3) for val in preds]
    for value in rounded_preds:
            # Convert the rounded value to a string
            worksheet.write(row, col + 2, value, decimal_format)
            row +=1
    row = 1
    for Data, Interval in zip(dataset.Data, dataset.Interval):
            worksheet.write(row, col + 0, Data, date_format)
            worksheet.write(row, col + 1, Interval)
            row +=1

    workbook.close()

#===============================================================================Forcasting RAAL Consumption=================================================================

def predicting_exporting_Consumption_RAAL(dataset):
    # Predict on forecast data
    forecast_dataset = dataset.copy()
    st.write(forecast_dataset)
    forecast_dataset["Month"] = forecast_dataset.Data.dt.month
    forecast_dataset["WeekDay"] = forecast_dataset.Data.dt.weekday
    forecast_dataset["Holiday"] = 0
    for holiday in forecast_dataset["Data"].unique():
        if holiday in holidays.ds.values:
            forecast_dataset["Holiday"][forecast_dataset["Data"] == holiday] = 1

    # Restructuring the dataset
    forecast_dataset = forecast_dataset[["WeekDay", "Month", "Holiday", "Interval", "Temperatura"]]
    # Loading the model
    xgb_loaded = joblib.load("./RAAL/Consumption/XGB_Consumption_RAAL.pkl")
    preds = xgb_loaded.predict(forecast_dataset.values)
    #Exporting Results to Excel
    workbook = xlsxwriter.Workbook('./RAAL/Consumption/Results/Results_Consumption_xgb_RAAL_WM.xlsx')
    worksheet = workbook.add_worksheet("Consumption_Predictions")
    date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
    # Define a format for cells with three decimal places
    decimal_format = workbook.add_format({'num_format': '0.000'})
    row = 1
    col = 0
    worksheet.write(0,0,"Data")
    worksheet.write(0,1,"Interval")
    worksheet.write(0,2,"Prediction")
    for value in preds:
        worksheet.write(row, col + 2, value, decimal_format)
        row +=1
    row = 1
    for Data, Interval in zip(dataset.Data, dataset.Interval):
        worksheet.write(row, col + 0, Data, date_format)
        worksheet.write(row, col + 1, Interval)
        row +=1

    workbook.close()

#================================================================================Rendering RAAL Consumption and Production pages=============================================

def render_RAAL_cons_prod(start_date, end_date):
    forecast = st.radio("Choose Forecast:", ["Consumption", "Production"])
    if forecast == "Consumption":
        if st.button("Run Consumption Forecast"):
            CEF = "RAAL"
            # 1. Building the Input file
            # Iterate from start_date to end_date, day by day
            current_date = start_date
            while current_date <= end_date:
                print(current_date.strftime('%Y-%m-%d'))
                # Fetching data for the date
                ghi_data = fetch_ghi_data_from_api(lat, lon, current_date.strftime('%Y-%m-%d'), api_key)
                # Processing GHI data
                output_file_path = './RAAL/Consumption/Input/GHI_{}.xlsx'.format(current_date.strftime('%Y-%m-%d'))  # Replace with your desired output file path
                ghi_json_to_excel(ghi_data, output_file_path)
                # Increment the date by one day
                current_date += timedelta(days=1)
            # Concatenating the GHI data into one file
            concatenate_ghi_data(CEF, forecast)
            # Formatting the date column
            ghi_file_path = "./RAAL/Consumption/Input/Concatenated_Hourly_GHI.xlsx"
            process_and_save_excel(ghi_file_path)
            # Adding the lookup column
            # add_lookup_column_GHI()
            file_path = "./RAAL/Consumption/Input/Concatenated_Hourly_GHI.xlsx"
            date_col_name = "date"
            interval_col_name = "hour"
            ghi_df = prepare_lookup_column(file_path, date_col_name, interval_col_name)

            # Fetching the weather data
            nr_days = ((end_date - start_date).days + 2)*24
            print(nr_days)
            temps_clouds_data = fetch_temps_clouds_data_from_api(lat,lon,api_key,nr_days)
            print(temps_clouds_data)
            # Saving the json file
            json_data = temps_clouds_data  # Replace with your JSON data
            folder_path = "./RAAL/Consumption/Input/"  # Replace with your folder path
            file_name = "weather.json"  # Replace with your desired file name
            save_json_to_file(json_data, folder_path, file_name)
            # Processing the weather data
            json_file_path = './RAAL/Consumption/Input/weather.json'  # Replace with your JSON file path
            output_file_path = './RAAL/Consumption/Input/weather.xlsx'  # Replace with your desired output file path
            flatten_json_to_excel(json_file_path, output_file_path)
            creating_weather_date_hour_columns(CEF, forecast)
            # change_date_format_weather_wb()
            # add_lookup_column_weather()
            file_path = "./RAAL/Consumption/Input/weather.xlsx"
            date_col_name = "Data"
            interval_col_name = "Interval"
            weather_df = prepare_lookup_column(file_path, date_col_name, interval_col_name)

            # Vuilding the Input file
            building_input_file(CEF, forecast)
            # add_lookup_column_Input()
            # add_lookup_column_input_xlsxwriter()
            file_path = "./RAAL/Consumption/Input_Consumption_RAAL.xlsx"
            date_col_name = "Data"
            interval_col_name = "Interval"
            input_df = prepare_lookup_column(file_path, date_col_name, interval_col_name)
            # lookup_ghi_values()
            # lookup_weather_values()
            # Lookuping the GHI values
            input_df = pd.read_excel("./RAAL/Consumption/Input_Consumption_RAAL.xlsx")
            ghi_df = pd.read_excel("./RAAL/Consumption/Input/Concatenated_Hourly_GHI.xlsx")
            lookup_ghi_values(CEF, input_df, ghi_df, forecast)
            # Lookuping the temperatures and clouds values
            weather_df = pd.read_excel("./RAAL/Consumption/Input/weather.xlsx")
            lookup_weather_values(CEF, input_df, weather_df, forecast)
            # Predicting the Production
            if os.path.exists("./RAAL/Consumption/Input_Consumption_RAAL.xlsx"):
                # If the file exists, show the button
                df = pd.read_excel("./RAAL/Consumption/Input_Consumption_RAAL.xlsx")
                predicting_exporting_Consumption_RAAL(df)
                file_path_results = './RAAL/Consumption/Results/Results_Consumption_xgb_RAAL_WM.xlsx'
                with open(file_path_results, "rb") as f:
                    excel_data = f.read()

                # Create a download link
                b64 = base64.b64encode(excel_data).decode()
                button_html = f"""
                     <a download="Consumption_Forecast_RAAL_WM_{end_date}.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
                     <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
                     </a> 
                     """
                st.markdown(button_html, unsafe_allow_html=True)
            else:
                # If the file does not exist, display a message
                st.error("Input file does not exist. Please ensure the file is in the correct location before proceeding.")
    else:
        if st.button("Run Production Forecast"):
            CEF = "RAAL"
            # 1. Building the Input file
            # Iterate from start_date to end_date, day by day
            current_date = start_date
            while current_date <= end_date:
                print(current_date.strftime('%Y-%m-%d'))
                # Fetching data for the date
                ghi_data = fetch_ghi_data_from_api(lat, lon, current_date.strftime('%Y-%m-%d'), api_key)
                # Processing GHI data
                output_file_path = './RAAL/Production/Input/GHI_{}.xlsx'.format(current_date.strftime('%Y-%m-%d'))  # Replace with your desired output file path
                ghi_json_to_excel(ghi_data, output_file_path)
                # Increment the date by one day
                current_date += timedelta(days=1)
            # Concatenating the GHI data into one file
            concatenate_ghi_data(CEF, forecast)
            # Formatting the date column
            ghi_file_path = "./RAAL/Production/Input/Concatenated_Hourly_GHI.xlsx"
            process_and_save_excel(ghi_file_path)
            # Adding the lookup column
            # add_lookup_column_GHI()
            file_path = "./RAAL/Production/Input/Concatenated_Hourly_GHI.xlsx"
            date_col_name = "date"
            interval_col_name = "hour"
            ghi_df = prepare_lookup_column(file_path, date_col_name, interval_col_name)

            # Fetching the weather data
            nr_days = ((end_date - start_date).days + 2)*24
            print(nr_days)
            temps_clouds_data = fetch_temps_clouds_data_from_api(lat,lon,api_key,nr_days)
            print(temps_clouds_data)
            # Saving the json file
            json_data = temps_clouds_data  # Replace with your JSON data
            folder_path = "./RAAL/Production/Input/"  # Replace with your folder path
            file_name = "weather.json"  # Replace with your desired file name
            save_json_to_file(json_data, folder_path, file_name)
            # Processing the weather data
            json_file_path = './RAAL/Production/Input/weather.json'  # Replace with your JSON file path
            output_file_path = './RAAL/Production/Input/weather.xlsx'  # Replace with your desired output file path
            flatten_json_to_excel(json_file_path, output_file_path)
            creating_weather_date_hour_columns(CEF, forecast)
            # change_date_format_weather_wb()
            # add_lookup_column_weather()
            file_path = "./RAAL/Production/Input/weather.xlsx"
            date_col_name = "Data"
            interval_col_name = "Interval"
            weather_df = prepare_lookup_column(file_path, date_col_name, interval_col_name)

            # Vuilding the Input file
            building_input_file(CEF, forecast)
            # add_lookup_column_Input()
            # add_lookup_column_input_xlsxwriter()
            file_path = "./RAAL/Production/Input_Production_RAAL.xlsx"
            date_col_name = "Data"
            interval_col_name = "Interval"
            input_df = prepare_lookup_column(file_path, date_col_name, interval_col_name)
            # lookup_ghi_values()
            # lookup_weather_values()
            # Lookuping the GHI values
            input_df = pd.read_excel("./RAAL/Production/Input_Production_RAAL.xlsx")
            ghi_df = pd.read_excel("./RAAL/Production/Input/Concatenated_Hourly_GHI.xlsx")
            lookup_ghi_values(CEF, input_df, ghi_df, forecast)
            # Lookupinh the temperatures and clouds values
            weather_df = pd.read_excel("./RAAL/Production/Input/weather.xlsx")
            lookup_weather_values(CEF, input_df, weather_df, forecast)
            # Predicting the Production
            if os.path.exists("./RAAL/Production/Input_Production_RAAL.xlsx"):
                # If the file exists, show the button
                df = pd.read_excel("./RAAL/Production/Input_Production_RAAL.xlsx")
                predicting_exporting_RAAL(df)
                file_path_results = './RAAL/Production/Results/Results_Production_xgb_RAAL_wm.xlsx'
                with open(file_path_results, "rb") as f:
                    excel_data = f.read()

                # Create a download link
                b64 = base64.b64encode(excel_data).decode()
                button_html = f"""
                     <a download="Production_Forecast_RAAL_WM_{current_date}.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
                     <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
                     </a> 
                     """
                st.markdown(button_html, unsafe_allow_html=True)
            else:
                # If the file does not exist, display a message
                st.error("Input file does not exist. Please ensure the file is in the correct location before proceeding.")

#===============================================================================Forcasting Solina Production=================================================================

def predicting_exporting_Solina(dataset):
    xgb_loaded = joblib.load("./Solina/Production/rs_xgb_Solina_prod_WM.pkl")
    dataset_forecast = dataset.copy()
    dataset.dropna(inplace=True)
    dataset_forecast = dataset_forecast[["Data", "Interval", "Radiatie", "Temperatura"]]
    dataset_forecast["Month"] = dataset_forecast.Data.dt.month
    dataset_forecast = dataset_forecast.drop("Data", axis=1)
    dataset_forecast.dropna(inplace=True)
    preds = xgb_loaded.predict(dataset_forecast.values)
    #Exporting Results to Excel
    workbook = xlsxwriter.Workbook('./Solina/Production/Results/Results_Production_xgb_Solina_wm.xlsx')
    worksheet = workbook.add_worksheet("Production_Predictions")
    date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
    # Define a format for cells with three decimal places
    # decimal_format = workbook.add_format({'num_format': '0.000'})
    row = 1
    col = 0
    worksheet.write(0,0,"Data")
    worksheet.write(0,1,"Interval")
    worksheet.write(0,2,"Prediction")
    # Rounded values
    rounded_preds = [round(val, 3) for val in preds]
    for value in rounded_preds:
            worksheet.write(row, col + 2, value)
            row +=1
    row = 1
    for Data, Interval in zip(dataset.Data, dataset.Interval):
            worksheet.write(row, col + 0, Data, date_format)
            worksheet.write(row, col + 1, Interval)
            row +=1

    workbook.close()

#===============================================================================Forcasting Solina Consumption=================================================================

def predicting_exporting_Consumption_Solina(forecast_dataset):
    # Predict on forecast data
    forecast_dataset["Month"] = forecast_dataset.Data.dt.month
    forecast_dataset["WeekDay"] = forecast_dataset.Data.dt.weekday
    forecast_dataset["Holiday"] = 0
    for holiday in forecast_dataset["Data"].unique():
            if holiday in holidays.ds.values:
                    forecast_dataset["Holiday"][forecast_dataset["Data"] == holiday] = 1

    # Restructuring the dataset
    forecast_dataset = forecast_dataset[["WeekDay", "Month", "Holiday", "Interval", "Temperatura"]]
    forecast_dataset.dropna(inplace=True)
    # Loading the model
    xgb_loaded = joblib.load("./Solina/Consumption/XGB_Consumption_Temperature.pkl")
    preds = xgb_loaded.predict(forecast_dataset.values)
    #Exporting Results to Excel
    workbook = xlsxwriter.Workbook('./Solina/Consumption/Results/Results_Consumption_xgb_Solina_WM.xlsx')
    worksheet = workbook.add_worksheet("Prediction_Consumption")
    # Define a format for cells with three decimal places
    decimal_format = workbook.add_format({'num_format': '0.000'})

    row = 1
    col = 0
    worksheet.write(0,0,"Prediction")
    # worksheet.write(0,1,"Real")

    for value in preds:
            worksheet.write(row, col, value, decimal_format)
            row +=1
    # row = 1
    # for value in y_test:
    #     worksheet.write(row, col + 1, value)
    #     row +=1

    workbook.close()

#==============================================================================Rendering Solina Consumption Production Page===================================================
def render_prod_cons_Solina_page(start_date, end_date):
    forecast = st.radio("Choose Forecast:", ["Consumption", "Production"])
    if forecast == "Production":
        if st.button("Run Production Forecast"):
            CEF = "Solina"
            # Iterate from start_date to end_date, day by day
            current_date = start_date
            while current_date <= end_date:
                print(current_date.strftime('%Y-%m-%d'))
                # Fetching data for the date
                ghi_data = fetch_ghi_data_from_api(lat, lon, current_date.strftime('%Y-%m-%d'), api_key)
                # Processing GHI data
                output_file_path = './Solina/Production/Input/GHI_{}.xlsx'.format(current_date.strftime('%Y-%m-%d'))  # Replace with your desired output file path
                ghi_json_to_excel(ghi_data, output_file_path)
                # Increment the date by one day
                current_date += timedelta(days=1)
            # Concatenating the GHI data into one file
            concatenate_ghi_data(CEF, forecast)
            # Formatting the date column
            ghi_file_path = "./Solina/Production/Input/Concatenated_Hourly_GHI.xlsx"
            process_and_save_excel(ghi_file_path)
            # Adding the lookup column
            # add_lookup_column_GHI()
            file_path = "./Solina/Production/Input/Concatenated_Hourly_GHI.xlsx"
            date_col_name = "date"
            interval_col_name = "hour"
            ghi_df = prepare_lookup_column(file_path, date_col_name, interval_col_name)

            # Fetching the weather data
            nr_days = ((end_date - start_date).days + 2)*24
            print(nr_days)
            temps_clouds_data = fetch_temps_clouds_data_from_api(lat,lon,api_key,nr_days)
            print(temps_clouds_data)
            # Saving the json file
            json_data = temps_clouds_data  # Replace with your JSON data
            folder_path = "./Solina/Production/Input/"  # Replace with your folder path
            file_name = "weather.json"  # Replace with your desired file name
            save_json_to_file(json_data, folder_path, file_name)
            # Processing the weather data
            json_file_path = './Solina/Production/Input/weather.json'  # Replace with your JSON file path
            output_file_path = './Solina/Production/Input/weather.xlsx'  # Replace with your desired output file path
            flatten_json_to_excel(json_file_path, output_file_path)
            creating_weather_date_hour_columns(CEF, forecast)
            # change_date_format_weather_wb()
            # add_lookup_column_weather()
            file_path = "./Solina/Production/Input/weather.xlsx"
            date_col_name = "Data"
            interval_col_name = "Interval"
            weather_df = prepare_lookup_column(file_path, date_col_name, interval_col_name)

            # Building the Input file
            building_input_file(CEF, forecast)
            # add_lookup_column_Input()
            # add_lookup_column_input_xlsxwriter()
            file_path = "./Solina/Production/Input_Production_Solina.xlsx"
            date_col_name = "Data"
            interval_col_name = "Interval"
            input_df = prepare_lookup_column(file_path, date_col_name, interval_col_name)
            # lookup_ghi_values()
            # lookup_weather_values()
            # Lookuping the GHI values
            input_df = pd.read_excel("./Solina/Production/Input_Production_Solina.xlsx")
            ghi_df = pd.read_excel("./Solina/Production/Input/Concatenated_Hourly_GHI.xlsx")
            lookup_ghi_values(CEF, input_df, ghi_df)
            # Lookupinh the temperatures and clouds values
            weather_df = pd.read_excel("./Solina/Production/Input/weather.xlsx")
            lookup_weather_values(CEF, input_df, weather_df)
            if os.path.exists("./Solina/Production/Input_Production_Solina.xlsx"):
            # Check if the file exists
                df = pd.read_excel("./Solina/Production/Input_Production_Solina.xlsx")
                predicting_exporting_Solina(df)
                file_path_results = './Solina/Production/Results/Results_Production_xgb_Solina_wm.xlsx'
                with open(file_path_results, "rb") as f:
                    excel_data = f.read()

                # Create a download link
                b64 = base64.b64encode(excel_data).decode()
                button_html = f"""
                     <a download="Production_Forecast_Solina_WM_{end_date}.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
                     <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
                     </a> 
                     """
                st.markdown(button_html, unsafe_allow_html=True)
            else:
                # If the file does not exist, display a message
                st.error("Input file does not exist. Please ensure the file is in the correct location before proceeding.")
    else:
        if st.button("Run Consumption Forecast"):
            CEF = "Solina"
            # Iterate from start_date to end_date, day by day
            current_date = start_date
            while current_date <= end_date:
                print(current_date.strftime('%Y-%m-%d'))
                # Fetching data for the date
                ghi_data = fetch_ghi_data_from_api(lat, lon, current_date.strftime('%Y-%m-%d'), api_key)
                # Processing GHI data
                output_file_path = './Solina/Consumption/Input/GHI_{}.xlsx'.format(current_date.strftime('%Y-%m-%d'))  # Replace with your desired output file path
                ghi_json_to_excel(ghi_data, output_file_path)
                # Increment the date by one day
                current_date += timedelta(days=1)
            # Concatenating the GHI data into one file
            concatenate_ghi_data(CEF, forecast)
            # Formatting the date column
            ghi_file_path = "./Solina/Consumption/Input/Concatenated_Hourly_GHI.xlsx"
            process_and_save_excel(ghi_file_path)
            # Adding the lookup column
            # add_lookup_column_GHI()
            file_path = "./Solina/Consumption/Input/Concatenated_Hourly_GHI.xlsx"
            date_col_name = "date"
            interval_col_name = "hour"
            ghi_df = prepare_lookup_column(file_path, date_col_name, interval_col_name)

            # Fetching the weather data
            nr_days = ((end_date - start_date).days + 2)*24
            print(nr_days)
            temps_clouds_data = fetch_temps_clouds_data_from_api(lat,lon,api_key,nr_days)
            print(temps_clouds_data)
            # Saving the json file
            json_data = temps_clouds_data  # Replace with your JSON data
            folder_path = "./Solina/Consumption/Input/"  # Replace with your folder path
            file_name = "weather.json"  # Replace with your desired file name
            save_json_to_file(json_data, folder_path, file_name)
            # Processing the weather data
            json_file_path = './Solina/Consumption/Input/weather.json'  # Replace with your JSON file path
            output_file_path = './Solina/Consumption/Input/weather.xlsx'  # Replace with your desired output file path
            flatten_json_to_excel(json_file_path, output_file_path)
            creating_weather_date_hour_columns(CEF, forecast)
            # change_date_format_weather_wb()
            # add_lookup_column_weather()
            file_path = "./Solina/Consumption/Input/weather.xlsx"
            date_col_name = "Data"
            interval_col_name = "Interval"
            weather_df = prepare_lookup_column(file_path, date_col_name, interval_col_name)

            # Building the Input file
            building_input_file(CEF, forecast)
            # add_lookup_column_Input()
            # add_lookup_column_input_xlsxwriter()
            file_path = "./Solina/Consumption/Input_Consumption_Solina.xlsx"
            date_col_name = "Data"
            interval_col_name = "Interval"
            input_df = prepare_lookup_column(file_path, date_col_name, interval_col_name)
            # lookup_ghi_values()
            # lookup_weather_values()
            # Lookuping the GHI values
            input_df = pd.read_excel("./Solina/Consumption/Input_Consumption_Solina.xlsx")
            ghi_df = pd.read_excel("./Solina/Consumption/Input/Concatenated_Hourly_GHI.xlsx")
            lookup_ghi_values(CEF, input_df, ghi_df, forecast)
            # Lookuping the temperatures and clouds values
            weather_df = pd.read_excel("./Solina/Consumption/Input/weather.xlsx")
            lookup_weather_values(CEF, input_df, weather_df, forecast)
            if os.path.exists("./Solina/Consumption/Input_Consumption_Solina.xlsx"):
            # Check if the file exists
                df = pd.read_excel("./Solina/Consumption/Input_Consumption_Solina.xlsx")
                predicting_exporting_Consumption_Solina(df)
                file_path_results = './Solina/Consumption/Results/Results_Consumption_xgb_Solina_WM.xlsx'
                with open(file_path_results, "rb") as f:
                    excel_data = f.read()

                # Create a download link
                b64 = base64.b64encode(excel_data).decode()
                button_html = f"""
                     <a download="Results_Consumption_Forecast_Solina_WM_{end_date}.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
                     <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
                     </a> 
                     """
                st.markdown(button_html, unsafe_allow_html=True)
            else:
                # If the file does not exist, display a message
                st.error("Input file does not exist. Please ensure the file is in the correct location before proceeding.")

solcast_api_key = os.getenv("solcast_api_key")
# output_path = "./Transavia/data/Bocsa.csv"
locations_cons = {"Lunca": {"lat": 46.427350, "lon": 23.905963}, "Brasov": {"lat": 45.642680, "lon": 25.588725},
                    "Santimbru": {"lat":46.135244, "lon":23.644428 }, "Bocsa": {"lat":45.377012 , "lon":21.718752}, "Cristian": {"lat":45.782114, "lon":24.029499},
                    "Cristuru": {"lat":46.292453, "lon":25.031714}, "Jebel": {"lat":45.562394 , "lon":21.214496}, "Medias": {"lat":46.157283, "lon":24.347167},
                    "Miercurea": {"lat":45.890054, "lon":23.791766}}

# Defining the fetching data function
def fetch_data(lat, lon, api_key, output_path):
    # Fetch data from the API
    api_url = "https://api.solcast.com.au/data/forecast/radiation_and_weather?latitude={}&longitude={}&hours=168&output_parameters=air_temp,dni,ghi&period=PT60M&format=csv&api_key={}".format(lat, lon, solcast_api_key)
    response = requests.get(api_url)
    print("Fetching data...")
    if response.status_code == 200:
        # Write the content to a CSV file
        with open(output_path, 'wb') as file:
            file.write(response.content)
    else:
        raise Exception(f"Failed to fetch data: Status code {response.status_code}")
#===============================================================================Rendering the Data Engineering page=================================================================

def render_data_eng_page():
    
    # Web App Title
    st.markdown('''
    # **Forecast Production**

    ''')

    # Streamlit interface
    location = st.radio(
    "Select location",
    ["Alba Iulia", "Prundu Bargaului"],
    captions = ["***Solina***", "***RAAL***"])

    if location == "Prundu Bargaului":
        lat = 47.2229
        lon = 24.7244
    else:
        lat = 46.073272
        lon = 23.580489

    # User inputs for start and end dates
    start_date = st.date_input("Start Date", datetime.today().date())
    end_date = st.date_input("End Date", datetime.today().date())
    
    if location == "Prundu Bargaului":
        render_RAAL_cons_prod(start_date, end_date)

    # Forecasting Solina Production
    elif location == "Alba Iulia":
        render_prod_cons_Solina_page(start_date, end_date)

    st.header("Fetching the Solcast data")
    if st.button("Fetch data"):
        # Iterating through the dictionary of PVPP locations
        for location in locations_cons.keys():
            print("Getting data for {}".format(location))
            output_path = f"./Transavia/data/{location}.csv"
            lat = locations_cons[location]["lat"]
            lon = locations_cons[location]["lon"]
            fetch_data(lat, lon, solcast_api_key, output_path)

        