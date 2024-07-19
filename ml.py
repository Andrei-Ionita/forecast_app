import streamlit as st
import streamlit.components.v1 as stc
import pandas as pd
import numpy as np
import base64
import xgboost as xgb
import joblib
import xlsxwriter
import os
import time
import zipfile
from datetime import datetime
import gdown
import requests
from openpyxl import load_workbook

session_start_time = time.time()

# Creating the holidays dataframe
# Creating the dictionary of holidays
New_year_and_day_after = pd.DataFrame({"holiday": "Anul Nou & A doua zi",
														"ds": pd.to_datetime(["2017-01-01", "2017-01-02", "2016-01-01", "2016-01-02", "2015-01-01", "2015-01-02", "2014-01-01", "2014-01-02", "2019-01-01",
																									"2019-01-02", "2018-01-01", "2018-01-02", "2020-01-01", "2020-01-02", "2021-01-01", "2021-01-02",
																									"2022-01-01", "2022-01-02", "2023-01-01", "2023-01-02","2024-01-01", "2024-01-02"]),
														"lower_window": -1,
														"upper_window": 1})	

National_holiday = pd.DataFrame({"holiday": "Ziua Nationala",
																 "ds": pd.to_datetime(["2016-12-01", "2015-12-01", "2014-12-01", "2018-12-01", "2019-12-01", "2020-12-01", "2021-12-01", "2022-12-01", "2023-12-01", "2024-12-01"]),
																 "lower_window": 0,
																 "upper_window": 1})
Ziua_Principatelor = pd.DataFrame({"holiday": "Ziua Principatelor",
																 "ds": pd.to_datetime(["2017-01-24", "2016-01-24", "2018-01-24", "2019-01-24", "2020-01-24", "2021-01-24", "2022-01-24", "2023-01-24", "2024-01-24"]),
																 "lower_window": 0,
																 "upper_window": 1})
Christmas = pd.DataFrame({"holiday": "Craciunul",
													"ds": pd.to_datetime(["2017-12-25", "2017-12-26", "2016-12-25", "2016-12-26", "2015-12-25", "2015-12-26", "2014-12-25", "2014-12-26", "2018-12-25", "2018-12-26", "2019-12-25", "2019-12-26", "2020-12-25", "2020-12-26", "2021-12-25", "2021-12-26",
																								"2022-12-25", "2022-12-26", "2023-12-25", "2023-12-26", "2024-12-25", "2024-12-26"]),
													"lower_window": -1,
													"upper_window": 1})
St_Andrew = pd.DataFrame({"holiday": "Sfantul Andrei",
													"ds": pd.to_datetime(["2017-11-30", "2016-11-30", "2015-11-30", "2014-11-30", "2018-11-30", "2019-11-30", "2020-11-30", "2021-11-30", "2022-11-30",
																								"2023-11-30", "2024-11-30"]),
													"lower_window": -1,
													"upper_window": 0})
Adormirea_Maicii_Domnului = pd.DataFrame({"holiday": "Adormirea Maicii Domnului",
																					"ds": pd.to_datetime(["2017-08-15", "2016-08-15", "2015-08-15", "2014-08-15", "2018-08-15", "2019-08-15", "2020-08-15", "2021-08-15","2022-08-15", "2024-08-15"])})
Rusalii = pd.DataFrame({"holiday": "Rusalii",
												"ds": pd.to_datetime(["2017-06-04", "2017-06-05", "2016-06-19", "2016-06-20", "2015-05-31", "2015-06-01", "2014-06-08", "2014-06-09", "2018-05-27", "2018-05-28", "2019-06-16", "2019-06-17", "2020-06-07", "2020-06-08", "2021-06-20", "2021-06-21",
																							"2022-06-12", "2022-06-13", "2023-06-04", "2023-06-05", "2024-06-24"])})
Ziua_Copilului = pd.DataFrame({"holiday": "Ziua Copilului",
															"ds": pd.to_datetime(["2017-06-01", "2018-06-01", "2019-06-01", "2020-06-01", "2021-06-01", "2022-06-01", "2023-06-01", "2024-06-01"])})
Ziua_Muncii = pd.DataFrame({"holiday": "Ziua Muncii",
														"ds": pd.to_datetime(["2017-05-01", "2016-05-01", "2015-05-01", "2014-05-01", "2018-05-01", "2019-05-01", "2020-05-01", "2021-05-01", "2022-05-01", "2023-05-01",
															"2024-05-01"])})
Pastele = pd.DataFrame({"holiday": "Pastele",
												"ds": pd.to_datetime(["2017-04-16", "2017-04-17", "2016-05-01", "2016-05-02", "2015-04-12", "2015-04-13", "2014-04-20", "2014-04-21", "2018-04-08", "2018-04-09", "2019-04-28", "2019-04-29", "2020-04-19", "2020-04-20", "2021-05-02", "2021-05-03",
																							"2022-04-24", "2022-04-25", "2023-04-16", "2023-04-17", "2024-05-06"]),
												"lower_window": -1,
												"upper_window": 1})
Vinerea_Mare = pd.DataFrame({"holiday": "Vinerea Mare",
														 "ds": pd.to_datetime(["2020-04-17", "2019-04-26", "2018-04-06", "2021-04-30", "2022-04-30", "2023-04-30", "2024-05-03"])})
Ziua_Unirii = pd.DataFrame({"holiday": "Ziua Unirii",
														"ds": pd.to_datetime(["2015-01-24", "2020-01-24", "2019-01-24", "2021-01-24", "2022-01-24", "2023-01-24", "2024-01-24"])})
Public_Holiday = pd.DataFrame({"holiday": "Public Holiday",
															"ds": pd.to_datetime(["2019-04-30"])})
holidays = pd.concat((New_year_and_day_after, National_holiday, Christmas, St_Andrew, Ziua_Principatelor, Adormirea_Maicii_Domnului, Rusalii, Ziua_Copilului, Ziua_Muncii,
											Pastele, Vinerea_Mare, Ziua_Unirii, Public_Holiday))

#=============================================================================Feetching the data for Transavia locations========================================================================
solcast_api_key = os.getenv("solcast_api_key")
# output_path = "./Transavia/data/Bocsa.csv"
# print("API Key:", solcast_api_key)  # Remove after debugging
if 'solcast_api_key' not in st.session_state:
	st.session_state['solcast_api_key'] = solcast_api_key = "pJFKpjXVuATTf72TB4hRc6lfu5W3Ww4_"


# Defining the fetching data function
def fetch_data(lat, lon, api_key, output_path):
	# Fetch data from the API
	api_url = "https://api.solcast.com.au/data/forecast/radiation_and_weather?latitude={}&longitude={}&hours=168&output_parameters=air_temp,cloud_opacity,ghi&period=PT60M&format=csv&api_key={}".format(lat, lon, solcast_api_key)
	response = requests.get(api_url)
	print("Fetching data...")
	if response.status_code == 200:
		# Write the content to a CSV file
		with open(output_path, 'wb') as file:
			file.write(response.content)
	else:
		print(response.text)  # Add this line to see the error message returned by the API
		raise Exception(f"Failed to fetch data: Status code {response.status_code}")

# ============================Crating the Input_production file==========
def creating_input_production_file(path):
	santimbru_data = pd.read_csv(f"{path}/Santimbru.csv")
	input_production = pd.read_excel("./Transavia/Production/Input_production.xlsx")
	input_production = input_production.copy()

	# Convert 'period_end' in santimbru to datetime
	santimbru_data['period_end'] = pd.to_datetime(santimbru_data['period_end'], errors='coerce')
	# Extract just the date part in the desired format (as strings)
	santimbru_dates = santimbru_data['period_end'].dt.strftime('%Y-%m-%d')

	# Write the dates from santimbru_dates to input_production.Data
	input_production['Data'] = santimbru_dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	input_production['Data'].fillna(method='bfill', inplace=True)

	# Completing the Interval column
	santimbru_intervals = santimbru_data["period_end"].dt.hour
	input_production["Interval"] = santimbru_intervals
	# Replace NaNs in the 'Interval' column with 0
	input_production['Interval'].fillna(0, inplace=True)

	# Completing the Radiatie column
	santimbru_radiatie = santimbru_data["ghi"]
	input_production["Radiatie"] = santimbru_radiatie

	# Completing the Temperatura column
	santimbru_temperatura = santimbru_data["air_temp"]
	input_production["Temperatura"] = santimbru_temperatura

	# Completing the Nori column
	santimbru_nori = santimbru_data["cloud_opacity"]
	input_production["Nori"] = santimbru_nori

	# Completing the Centrala column
	input_production["Centrala"] = "Abator_Oiejdea"

	# Copying the data for FNC PVPP
	copy_df = input_production.copy()
	copy_df['Centrala'] = "FNC"

	# Append the copied dataframe to the original dataframe
	input_production = pd.concat([input_production, copy_df])

	# Copying the data for F4 PVPP
	copy_df = input_production[input_production["Centrala"] == "Abator_Oiejdea"].copy()
	copy_df['Centrala'] = "F4"

	# Append the copied dataframe to the original dataframe
	input_production = pd.concat([input_production, copy_df])

	# Copying the data for Ciugud PVPP
	copy_df = input_production[input_production["Centrala"] == "Abator_Oiejdea"].copy()
	copy_df['Centrala'] = "Ciugud"

	# Append the copied dataframe to the original dataframe
	input_production = pd.concat([input_production, copy_df])

	# Copying the data for Abator Bocsa PVPP
	copy_df = input_production[input_production["Centrala"] == "Abator_Oiejdea"].copy()
	copy_df['Centrala'] = "Abator_Bocsa"
	# Append the copied dataframe to the original dataframe
	input_production = pd.concat([input_production, copy_df])

	bocsa_data = pd.read_csv(f"{path}/Bocsa.csv")

	# Completing the Radiatie column for Abator Bocsa
	bocsa_radiatie = bocsa_data["ghi"]
	input_production["Radiatie"][input_production["Centrala"] == "Abator_Bocsa"] = bocsa_radiatie

	# Completing the Temperatura column for Abator Bocsa
	bocsa_temperatura = bocsa_data["air_temp"]
	input_production["Temperatura"][input_production["Centrala"] == "Abator_Bocsa"] = bocsa_temperatura

	# Completing the Nori column for Abator Bocsa
	bocsa_nori = bocsa_data["cloud_opacity"]
	input_production["Nori"][input_production["Centrala"] == "Abator_Bocsa"] = bocsa_nori

	# Copying the data for Lunca PVPP
	copy_df = input_production[input_production["Centrala"] == "Abator_Oiejdea"].copy()
	copy_df['Centrala'] = "F24"
	# Append the copied dataframe to the original dataframe
	input_production = pd.concat([input_production, copy_df])

	lunca_data = pd.read_csv(f"{path}/Lunca.csv")

	# Completing the Radiatie column for F24
	lunca_radiatie = lunca_data["ghi"]
	input_production["Radiatie"][input_production["Centrala"] == "F24"] = lunca_radiatie

	# Completing the Temperatura column for F24
	lunca_temperatura = lunca_data["air_temp"]
	input_production["Temperatura"][input_production["Centrala"] == "F24"] = lunca_temperatura

	# Completing the Nori column for F24
	lunca_nori = lunca_data["cloud_opacity"]
	input_production["Nori"][input_production["Centrala"] == "F24"] = lunca_nori

	# Copying the data for Brasov PVPP
	copy_df = input_production[input_production["Centrala"] == "Abator_Oiejdea"].copy()
	copy_df['Centrala'] = "Brasov"
	# Append the copied dataframe to the original dataframe
	input_production = pd.concat([input_production, copy_df])

	brasov_data = pd.read_csv(f"{path}/Brasov.csv")

	# Completing the Radiatie column for Brasov
	brasov_radiatie = brasov_data["ghi"]
	input_production["Radiatie"][input_production["Centrala"] == "Brasov"] = brasov_radiatie

	# Completing the Temperatura column for Brasov
	brasov_temperatura = brasov_data["air_temp"]
	input_production["Temperatura"][input_production["Centrala"] == "Brasov"] = brasov_temperatura

	# Completing the Nori column for Brasov
	brasov_nori = brasov_data["cloud_opacity"]
	input_production["Nori"][input_production["Centrala"] == "Brasov"] = brasov_nori

	# Saving input_production to Excel
	input_production.to_excel("./Transavia/Production/Input_production_filled.xlsx", index=False)

# ===================================================================================TRANSAVIA FORECAST==================================================================================================================
def cleaning_input_files():
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx")
	input_santimbru = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	input_brasov[:] = ""
	input_santimbru[:] = ""
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)
	input_santimbru.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

# Creating the Input file for Santimbru region===========================
def creating_input_consumption_Santimbru():
	input = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	santimbru_data = pd.read_csv("./Transavia/data/Santimbru.csv")
	# Convert 'period_end' in santimbru to datetime
	santimbru_data['period_end'] = pd.to_datetime(santimbru_data['period_end'], errors='coerce')
	# Extract just the date part in the desired format (as strings)
	santimbru_dates = santimbru_data['period_end'].dt.strftime('%Y-%m-%d')

	# Write the dates from santimbru_dates to input_production.Data
	input['Data'] = santimbru_dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	input['Data'].fillna(method='bfill', inplace=True)
	# Completing the Interval column
	santimbru_intervals = santimbru_data["period_end"].dt.hour
	input["Interval"] = santimbru_intervals
	# Replace NaNs in the 'Interval' column with 0
	input['Interval'].fillna(0, inplace=True)
	# Completing the Temperatura column
	santimbru_temperatura = santimbru_data["air_temp"]
	input["Temperatura"] = santimbru_temperatura
	# Filling the IBD column
	input["IBD"] = "Abator"
	# Filling the Locatie column
	input["Locatie"] = "Santimbru"
	# Filling the PVPP column
	input["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	input["Data"] = pd.to_datetime(input["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	input['Lookup'] = input["Data"].dt.strftime('%d.%m.%Y') + str("F2")
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	main_df.to_excel('./Transavia/Consumption/Input_flow_Chicks.xlsx', index=False)

	# Filling the data for Ciugud
	ciugud_data = input[input["IBD"] == "Abator"].copy()
	ciugud_data["IBD"] = "Ciugud"
	ciugud_data["Flow_Chicks"] = ""
	ciugud_data["Lookup"] = ""
	input = pd.concat([input, ciugud_data])

	# Filling the data for F20-F21
	f20_f21_data = input[input["IBD"] == "Abator"].copy()
	f20_f21_data["IBD"] = "F20-F21"
	f20_f21_data["PVPP"] = ""
	# Filling the Flow_Chicks column for F20-F21
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f20_f21_data['Lookup'] = f20_f21_data["Data"].dt.strftime('%d.%m.%Y') + str("F20")
	input = pd.concat([input, f20_f21_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = main_df.copy()
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F3
	f3_data = input[input["IBD"] == "Abator"].copy()
	f3_data["IBD"] = "F3"
	f3_data["PVPP"] = ""
	# Filling the Flow_Chicks column for F3
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f3_data['Lookup'] = f3_data["Data"].dt.strftime('%d.%m.%Y') + str("F3")
	input = pd.concat([input, f3_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = main_df.copy()
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F4
	f4_data = input[input["IBD"] == "Abator"].copy()
	f4_data["IBD"] = "F4"
	# Filling the Flow_Chicks column for F4
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f4_data['Lookup'] = f4_data["Data"].dt.strftime('%d.%m.%Y') + str("F4")
	input = pd.concat([input, f4_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = main_df.copy()
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F5
	f5_data = input[input["IBD"] == "Abator"].copy()
	f5_data["IBD"] = "F5"
	# Filling the Flow_Chicks column for F5
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f5_data['Lookup'] = f5_data["Data"].dt.strftime('%d.%m.%Y') + str("F5")
	input = pd.concat([input, f5_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = main_df.copy()
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F7
	f7_data = input[input["IBD"] == "Abator"].copy()
	f7_data["IBD"] = "F7"
	f7_data["PVPP"] = ""
	# Filling the Flow_Chicks column for F7
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f7_data['Lookup'] = f7_data["Data"].dt.strftime('%d.%m.%Y') + str("F7")
	input = pd.concat([input, f7_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = main_df.copy()
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for FNC
	fnc_data = input[input["IBD"] == "Abator"].copy()
	fnc_data["IBD"] = "FNC"
	# Filling the Flow_Chicks column for FNC
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	fnc_data['Lookup'] = fnc_data["Data"].dt.strftime('%d.%m.%Y') + str("FNC")
	input = pd.concat([input, fnc_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = main_df.copy()
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for Abator Bocsa
	bocsa_data = pd.read_csv("./Transavia/data/Bocsa.csv")
	abator_bocsa_data = input[input["IBD"] == "Abator"].copy()
	abator_bocsa_data["IBD"] = "Abator_Bocsa"
	abator_bocsa_data["Locatie"] = "Bocsa"
	# Completing the Temperatura column
	bocsa_temperatura = bocsa_data["air_temp"]
	abator_bocsa_data["Temperatura"] = bocsa_temperatura
	# Completing the Radiatie column
	# bocsa_ghi = bocsa_data["ghi"]
	# abator_bocsa_data["Radiatie"] = bocsa_ghi
	# Filling the Flow_Chicks column for Abator Bocsa
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	abator_bocsa_data['Lookup'] = abator_bocsa_data["Data"].dt.strftime('%d.%m.%Y') + str("F15S1")
	abator_bocsa_data['Lookup2'] = abator_bocsa_data["Data"].dt.strftime('%d.%m.%Y') + str("F22")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Bocsa')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = abator_bocsa_data
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict) + main_df['Lookup2'].map(lookup_dict)
	input = pd.concat([input, abator_bocsa_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for Ferma_Bocsa
	ferma_bocsa_data = input[input["IBD"] == "Abator_Bocsa"].copy()
	ferma_bocsa_data["IBD"] = "Ferma_Bocsa"
	# Filling the Flow_Chicks column for Ferma Bocsa
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	ferma_bocsa_data['Lookup'] = ferma_bocsa_data["Data"].dt.strftime('%d.%m.%Y') + str("F15S2")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Bocsa')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = ferma_bocsa_data
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = pd.concat([input, ferma_bocsa_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F_Cristian
	cristian_data = pd.read_csv("./Transavia/data/Cristian.csv")
	f_cristian_data = input[input["IBD"] == "Abator"].copy()
	f_cristian_data["IBD"] = "F_Cristian"
	f_cristian_data["Locatie"] = "Cristian"
	f_cristian_data["PVPP"] = ""
	f_cristian_data["Flow_Chicks"] = ""
	f_cristian_data["Lookup"] = ""
	# Completing the Temperatura column
	cristian_temperatura = cristian_data["air_temp"]
	f_cristian_data["Temperatura"] = cristian_temperatura
	# Completing the Radiatie column
	cristian_ghi = cristian_data["ghi"]
	cristian_data["Radiatie"] = ""
	input = pd.concat([input, f_cristian_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F10
	cristuru_data = pd.read_csv("./Transavia/data/Cristuru.csv")
	f10_data = input[input["IBD"] == "Abator"].copy()
	f10_data["IBD"] = "F10"
	f10_data["Locatie"] = "Cristuru Secuiesc"
	f10_data["PVPP"] = ""
	# Completing the Temperatura column
	cristuru_temperatura = cristuru_data["air_temp"]
	f10_data["Temperatura"] = cristuru_temperatura
	# Filling the Flow_Chicks column for F10
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f10_data['Lookup'] = f10_data["Data"].dt.strftime('%d.%m.%Y') + str("F10")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = f10_data
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = pd.concat([input, f10_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for Jebel1
	jebel_data = pd.read_csv("./Transavia/data/Jebel.csv")
	jebel1_data = input[input["IBD"] == "Abator"].copy()
	jebel1_data["IBD"] = "Jebel1"
	jebel1_data["Locatie"] = "Jebel"
	jebel1_data["PVPP"] = ""
	jebel1_data["Flow_Chicks"] = ""
	jebel1_data["Lookup"] = ""
	# Completing the Temperatura column
	jebel_temperatura = jebel_data["air_temp"]
	jebel1_data["Temperatura"] = jebel_temperatura
	input = pd.concat([input, jebel1_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F6
	lunca_data = pd.read_csv("./Transavia/data/Lunca.csv")
	f6_data = input[input["IBD"] == "Abator"].copy()
	f6_data["IBD"] = "F6"
	f6_data["Locatie"] = "Lunca Muresului"
	f6_data["PVPP"] = ""
	# Completing the Temperatura column
	lunca_temperatura = lunca_data["air_temp"]
	f6_data["Temperatura"] = lunca_temperatura
	# Filling the Flow_Chicks column for F6
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f6_data['Lookup'] = f6_data["Data"].dt.strftime('%d.%m.%Y') + str("F6")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = f6_data
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = pd.concat([input, f6_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F17
	medias_data = pd.read_csv("./Transavia/data/Medias.csv")
	f17_data = input[input["IBD"] == "Abator"].copy()
	f17_data["IBD"] = "F17"
	f17_data["Locatie"] = "Medias"
	f17_data["PVPP"] = ""
	f17_data["Flow_Chicks"] = ""
	f17_data["Lookup"] = ""
	# Completing the Temperatura column
	medias_temperatura = medias_data["air_temp"]
	f17_data["Temperatura"] = medias_temperatura
	input = pd.concat([input, f17_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F9
	miercurea_data = pd.read_csv("./Transavia/data/Miercurea.csv")
	f9_data = input[input["IBD"] == "Abator"].copy()
	f9_data["IBD"] = "F9"
	f9_data["Locatie"] = "Miercurea Sibiului"
	f9_data["PVPP"] = ""
	# Completing the Temperatura column
	miercurea_temperatura = miercurea_data["air_temp"]
	f9_data["Temperatura"] = miercurea_temperatura
	# Filling the Flow_Chicks column for F9
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f9_data['Lookup'] = f9_data["Data"].dt.strftime('%d.%m.%Y') + str("F9")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = f9_data
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = pd.concat([input, f9_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)

	# Filling the data for F8
	lunca_data = pd.read_csv("./Transavia/data/Lunca.csv")
	f8_data = input[input["IBD"] == "Abator"].copy()
	f8_data["IBD"] = "F8"
	f8_data["Locatie"] = "Lunca Muresului"
	f8_data["PVPP"] = ""
	# Completing the Temperatura column
	lunca_temperatura = lunca_data["air_temp"]
	f8_data["Temperatura"] = lunca_temperatura
	# Filling the Flow_Chicks column for F8
	# Adding the Lookup column to the Input.xlsx file
	# Create the 'Lookup' column by concatenating the 'Data' and 'IBD' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	f8_data['Lookup'] = f8_data["Data"].dt.strftime('%d.%m.%Y') + str("F8")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Alba')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = f8_data
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Temperatures values
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	input = pd.concat([input, f8_data])
	input.to_excel("./Transavia/Consumption/Input/Input.xlsx", index=False)


	return input

# Creating the Input file for Brasov region====================================================
def creating_input_cons_file_Brasov():
	# Filling the data for 594020100002383007
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx")
	brasov_data = pd.read_csv("./Transavia/data/Brasov.csv")
	# Convert 'period_end' in santimbru to datetime
	brasov_data['period_end'] = pd.to_datetime(brasov_data['period_end'], errors='coerce')
	# Extract just the date part in the desired format (as strings)
	brasov_dates = brasov_data['period_end'].dt.strftime('%Y-%m-%d')

	# Write the dates from santimbru_dates to input_production.Data
	input_brasov['Data'] = brasov_dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	input_brasov['Data'].fillna(method='bfill', inplace=True)
	# Completing the Interval column
	brasov_intervals = brasov_data["period_end"].dt.hour
	input_brasov["Interval"] = brasov_intervals
	# Replace NaNs in the 'Interval' column with 0
	input_brasov['Interval'].fillna(0, inplace=True)
	# Completing the Temperatura column
	brasov_temperatura = brasov_data["air_temp"]
	input_brasov["Temperatura"] = brasov_temperatura
	# Filling the IBD column
	input_brasov["POD"] = "594020100002383007"
	# Filling the PVPP column
	input_brasov["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	input_brasov["Data"] = pd.to_datetime(input_brasov["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	input_brasov['Lookup'] = input_brasov["Data"].dt.strftime('%d.%m.%Y') + str("F25")
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	main_df = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx")
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	main_df['Flow_Chicks'] = main_df['Lookup'].map(lookup_dict)
	main_df.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Preserving the POD column format
	from openpyxl import Workbook
	from openpyxl.utils.dataframe import dataframe_to_rows

	wb = Workbook()
	ws = wb.active

	for r in dataframe_to_rows(main_df, index=False, header=True):
		ws.append(r)

	# Assuming the 'POD' column is the first column, starting from the second row
	for cell in ws['E'][1:]:  # Skip header row
		cell.value = str(cell.value)  # Reinforce string format
		cell.number_format = '@'

	wb.save("./Transavia/Consumption/output_formatted.xlsx")

	# Filling the data for 594020100002224041
	input_brasov = pd.read_excel("./Transavia/Consumption/output_formatted.xlsx")
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002383007"].copy()
	new_data["POD"] = "594020100002224041"
	new_data["PVPP"] = ""
	new_data["Flow_Chicks"] = ""
	new_data["Lookup"] = ""
	print(new_data)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)

	# Filling the data for 594020100002273568
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002273568"
	new_data["PVPP"] = ""
	new_data["Flow_Chicks"] = ""
	new_data["Lookup"] = ""
	print(new_data)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)

	# Filling the data for 594020100002382970
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002382970"
	new_data["PVPP"] = 1
	new_data["Flow_Chicks"] = ""
	new_data["Lookup"] = ""
	print(new_data)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)

	# Filling the data for 594020100002383014
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002383014"
	# Filling the PVPP column
	new_data["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F30/S1")
	new_data['Lookup2'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F30/S2")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict) + new_data['Lookup2'].map(lookup_dict)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020100002383069
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002383069"
	new_data["PVPP"] = 1
	new_data["Flow_Chicks"] = ""
	new_data["Lookup"] = ""
	print(new_data)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)

	# Filling the data for 594020100002383502
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002383502"
	# Filling the PVPP column
	new_data["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F26")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020100002383519
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002383519"
	# Filling the PVPP column
	new_data["PVPP"] = ""
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F27")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020100002384233
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002384233"
	new_data["PVPP"] = ""
	new_data["Flow_Chicks"] = ""
	new_data["Lookup"] = ""
	print(new_data)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)

	# Filling the data for 594020100002384691
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002384691"
	# Filling the PVPP column
	new_data["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F31")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020100002836497
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002836497"
	# Filling the PVPP column
	new_data["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F27")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020100002841279
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002841279"
	# Filling the PVPP column
	new_data["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F29")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020100002967269
	input_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx").copy()
	print(input_brasov)
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020100002967269"
	# Filling the PVPP column
	new_data["PVPP"] = 1
	# Filling the Flow_Chicks column
	# Adding the Lookup column to the Input.xlsx file
	# Ensure the 'Data' column is in datetime format
	new_data["Data"] = pd.to_datetime(new_data["Data"])
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	new_data['Lookup'] = new_data["Data"].dt.strftime('%d.%m.%Y') + str("F28")
	# Adding the Lookup column to the Fluxuri_pui.xlsx file
	df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsm", sheet_name='Brasov')

	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])

	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Loc"].astype(str)
	df.to_excel("./Transavia/Consumption/Fluxuri_pui.xlsx", index=False)
	# Mapping the Flow_Chicks column fo the input
	lookup_df = pd.read_excel("./Transavia/Consumption/Fluxuri_pui.xlsx")
	# Create a dictionary from lookup_df for efficient lookup
	lookup_dict = lookup_df.set_index("Lookup")["Fluxuri_Input"].to_dict()
	# Perform the lookup by mapping the 'Lookup' column in main_df to the values in lookup_dict
	new_data['Flow_Chicks'] = new_data['Lookup'].map(lookup_dict)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel('./Transavia/Consumption/Input/Input_Brasov.xlsx', index=False)

	# Filling the data for 594020300002359730
	input_brasov["POD"] = input_brasov["POD"].astype(str)
	new_data = input_brasov[input_brasov["POD"] == "594020100002224041"].copy()
	new_data["POD"] = "594020300002359730"
	new_data["PVPP"] = ""
	new_data["Flow_Chicks"] = ""
	new_data["Lookup"] = ""
	print(new_data)
	input_brasov = pd.concat([input_brasov, new_data])
	input_brasov.to_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx", index=False)

def predicting_exporting_Transavia(dataset):
	datasets_forecast = dataset.copy()
	CEFs = datasets_forecast.Centrala.unique()
	dataset_forecast = {elem : pd.DataFrame for elem in CEFs}
	for CEF in CEFs:
		print("Predicting for {}".format(CEF))
		xgb_loaded = joblib.load("./Transavia/Production/Models/rs_xgb_{}.pkl".format(CEF))
		dataset_forecast = datasets_forecast[:][datasets_forecast.Centrala == CEF]
		dataset_forecast["Data"] = pd.to_datetime(dataset_forecast["Data"])
		dataset_forecast["Month"] = dataset_forecast.Data.dt.month
		if CEF in ["F24"]:
			df_forecast = dataset_forecast.drop(["Data", "Nori", "Centrala"], axis=1)
		else:
			df_forecast = dataset_forecast.drop(["Data", "Centrala"], axis=1)
		preds = xgb_loaded.predict(df_forecast.values)
		# Exporting Results to Excel
		workbook = xlsxwriter.Workbook("./Transavia/Production/Results/Results_daily_{}.xlsx".format(CEF))
		worksheet = workbook.add_worksheet("Prediction_Production")

		worksheet.write(0,0,"Data")
		worksheet.write(0,1,"Interval")
		worksheet.write(0,2,"Production")
		date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
		row = 1
		col = 0
		for value in preds:
		  worksheet.write(row, col+2, value)
		  row +=1
		row = row - len(preds)
		for data in dataset_forecast["Data"]:
		  worksheet.write(row, col, data, date_format)
		  row +=1
		row = row - len(dataset_forecast["Data"])
		for value in dataset_forecast["Interval"]:
		  worksheet.write(row, col+1, value)
		  row +=1
		workbook.close()

def predicting_exporting_consumption_Santimbru(dataset):
	# Importing the dataset
	dataset_forecast = dataset
	IBDs = dataset_forecast.IBD.unique()
	IBDs_PVPP = ["Abator", "Abator_Bocsa", "Ciugud", "F5", "Ferma_Bocsa", "FNC", "F4"]
	datasets_forecast = {elem : pd.DataFrame for elem in IBDs}
	for IBD in IBDs:
		datasets_forecast[IBD] = dataset_forecast[:][dataset_forecast.IBD == IBD]
		datasets_forecast[IBD]["WeekDay"] = datasets_forecast[IBD].Data.dt.weekday
		datasets_forecast[IBD]["Month"] = datasets_forecast[IBD].Data.dt.month
		datasets_forecast[IBD]["Holiday"] = 0
		for holiday in datasets_forecast[IBD]["Data"].unique():
			if holiday in holidays.ds.values:
				datasets_forecast[IBD]["Holiday"][datasets_forecast[IBD]["Data"] == holiday] = 1
		if len(datasets_forecast[IBD]["Flow_Chicks"].value_counts()) > 0 and len(datasets_forecast[IBD]["Radiatie"].value_counts()) > 0:
			## Restructuring the dataset
			datasets_forecast[IBD] = datasets_forecast[IBD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Flow_Chicks", "Radiatie"]]
		elif len(datasets_forecast[IBD]["Flow_Chicks"].value_counts()) > 0:
			datasets_forecast[IBD] = datasets_forecast[IBD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Flow_Chicks"]]
		elif len(datasets_forecast[IBD]["Radiatie"].value_counts()) > 0:
			datasets_forecast[IBD] = datasets_forecast[IBD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Radiatie"]]
		else:
		  datasets_forecast[IBD] = datasets_forecast[IBD][["Month", "WeekDay","Holiday", "Interval", "Temperatura"]]
	# datasets_forecast[IBD].replace([np.inf, -np.inf], np.nan)
	# datasets_forecast[IBD].dropna(inplace = True)
		# Check if the cons place has PVPP and add the column, if it does
		if IBD in IBDs_PVPP:
			datasets_forecast[IBD]["PVPP"] = 1
	# Predicting
	predictions = {}
	for IBD in datasets_forecast.keys():
		if IBD not in IBDs_PVPP:
			xgb_loaded = joblib.load("./Transavia/Consumption/Models_PVPP/rs_xgb_{}.pkl".format(IBD))
			print("Predicting for {}".format(IBD))
			# st.write(datasets_forecast[IBD])
			xgb_preds = xgb_loaded.predict(datasets_forecast[IBD].values)
			predictions[IBD] = xgb_preds
			predictions["Data"] = dataset_forecast["Data"]
			predictions["Interval"] = dataset_forecast["Interval"]
	# Predicting with PVPP models
	predictions_PVPP = {}
	for IBD in datasets_forecast.keys():
		if IBD in IBDs_PVPP:
			if os.path.isfile("./Transavia/Consumption/Models_PVPP/rs_xgb_{}_PVPP.pkl".format(IBD)):
				xgb_loaded = joblib.load("./Transavia/Consumption/Models_PVPP/rs_xgb_{}_PVPP.pkl".format(IBD))
				print("Predicting for {}".format(IBD))
				# st.write(datasets_forecast[IBD])
				xgb_preds = xgb_loaded.predict(datasets_forecast[IBD].values)
				predictions_PVPP[IBD] = xgb_preds
				predictions_PVPP["Data"] = dataset_forecast["Data"]
				predictions_PVPP["Interval"] = dataset_forecast["Interval"]
	# Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Transavia/Consumption/Results/XGB/Results_IBDs_daily.xlsx")
	worksheet = workbook.add_worksheet("Prediction_Consumption")

	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")
	worksheet.write(0,3,"IBD")
	worksheet.write(0,4,"Lookup")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	row = 1
	col = 0
	for IBD in datasets_forecast.keys():
		if IBD in predictions_PVPP.keys():
			for value in predictions_PVPP[IBD]:
				worksheet.write(row, col+2, value)
				worksheet.write(row, col+3, IBD)
				row +=1
		else:
			for value in predictions[IBD]:
				worksheet.write(row, col+2, value)
				worksheet.write(row, col+3, IBD)
				row +=1
	row = row - len(predictions[IBD])*(len(datasets_forecast.keys()))
	for data in dataset_forecast["Data"]:
		worksheet.write(row, col, datetime.date(data),date_format)
		worksheet.write_formula(row, col+4, "=A"+ str(row+1)+ "&" + "B"+ str(row+1)+ "&" + "D" + str(row+1))
		row +=1
	row = row - len(predictions["Data"])
	for interval in predictions["Interval"]:
		worksheet.write(row, col+1, interval)
		row +=1
		# row = 1
		# for value in y_test:
		#     worksheet.write(row, col + 1, value)
		#     row +=1
	workbook.close()

def predicting_exporting_consumption_Brasov(dataset):
	# Importing the dataset
	dataset_forecast = dataset
	PODs = dataset_forecast.POD.unique()
	datasets_forecast = {elem : pd.DataFrame for elem in PODs}
	for POD in PODs:
		datasets_forecast[POD] = dataset_forecast[:][dataset_forecast.POD == POD]
		datasets_forecast[POD]["WeekDay"] = datasets_forecast[POD].Data.dt.weekday
		datasets_forecast[POD]["Month"] = datasets_forecast[POD].Data.dt.month
		datasets_forecast[POD]["Holiday"] = 0
		for holiday in datasets_forecast[POD]["Data"].unique():
		  if holiday in holidays.ds.values:
			  datasets_forecast[POD]["Holiday"][datasets_forecast[POD]["Data"] == holiday] = 1
		if len(datasets_forecast[POD]["Flow_Chicks"].value_counts()) > 0 and len(datasets_forecast[POD]["PVPP"].value_counts()) > 0:
			datasets_forecast[POD] = datasets_forecast[POD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Flow_Chicks", "PVPP"]]
		elif len(datasets_forecast[POD]["Flow_Chicks"].value_counts()) > 0:
			datasets_forecast[POD] = datasets_forecast[POD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "Flow_Chicks"]]
		elif len(datasets_forecast[POD]["PVPP"].value_counts()) > 0:
			datasets_forecast[POD] = datasets_forecast[POD][["Month", "WeekDay","Holiday", "Interval", "Temperatura", "PVPP"]]
		else:
			datasets_forecast[POD] = datasets_forecast[POD][["Month", "WeekDay","Holiday", "Interval", "Temperatura"]]
		# datasets_forecast[POD].replace([np.inf, -np.inf], np.nan)
		# datasets_forecast[POD].dropna(inplace = True)
	
	# Predicting noPVPP
	predictions = {}
	for POD in datasets_forecast.keys():
		if "PVPP" in datasets_forecast[POD].columns:
			xgb_loaded = joblib.load("./Transavia/Consumption/Models_PVPP/Brasov_Models/rs_xgb_{}.pkl".format(POD))
			print("Predicting for {}".format(POD))
			# st.write(datasets_forecast[POD])
			xgb_preds = xgb_loaded.predict(datasets_forecast[POD].values)
			predictions[POD] = xgb_preds
			predictions["Data"] = dataset_forecast["Data"]
			predictions["Interval"] = dataset_forecast["Interval"]
		else:
			xgb_loaded = joblib.load("./Transavia/Consumption/Models_PVPP/Brasov_Models/rs_xgb_{}.pkl".format(POD))
			print("Predicting for without PVPP{}".format(POD))
			# st.write(datasets_forecast[POD])
			xgb_preds = xgb_loaded.predict(datasets_forecast[POD].values)
			predictions[POD] = xgb_preds
			predictions["Data"] = dataset_forecast["Data"]
			predictions["Interval"] = dataset_forecast["Interval"]
	# Predicting with PVPP
	predictions_PVPP = {}
	for POD in datasets_forecast.keys():
		if os.path.isfile(".Transavia/Consumption/Brasov_Models/rs_xgb_{}_PVPP.pkl".format(POD)):
		  xgb_loaded = joblib.load("./Transavia/Consumption/Brasov_Models/rs_xgb_{}_PVPP.pkl".format(POD))
		  print("Predicting for {}".format(POD))
		  # print(datasets_forecast[POD])
		  xgb_preds = xgb_loaded.predict(datasets_forecast[POD].values)
		  predictions_PVPP[POD] = xgb_preds
		  predictions_PVPP["Data"] = dataset_forecast["Data"]
		  predictions_PVPP["Interval"] = dataset_forecast["Interval"]

	# Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Transavia/Consumption/Results/XGB/Results_IBDs_daily_Brasov.xlsx")
	worksheet = workbook.add_worksheet("Prediction_Consumption")

	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")
	worksheet.write(0,3,"POD")
	worksheet.write(0,4,"Lookup")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	row = 1
	col = 0
	for POD in datasets_forecast.keys():
		if POD in predictions_PVPP.keys():
			for value in predictions_PVPP[POD]:
				worksheet.write(row, col+2, abs(value))
				worksheet.write(row, col+3, str(POD))
				row +=1
		else:
			for value in predictions[POD]:
				worksheet.write(row, col+2, abs(value))
				worksheet.write(row, col+3, str(POD))
				row +=1
	row = row - len(predictions[POD])*(len(datasets_forecast.keys()))
	for data in dataset_forecast["Data"]:
		worksheet.write(row, col, datetime.date(data),date_format)
		worksheet.write_formula(row, col+4, "=A"+ str(row+1)+ "&" + "B"+ str(row+1)+ "&" + "D" + str(row+1))
		row +=1
	row = row - len(predictions["Data"])
	for interval in predictions["Interval"]:
		worksheet.write(row, col+1, interval)
		row +=1
		# row = 1
		# for value in y_test:
		#     worksheet.write(row, col + 1, value)
		#     row +=1
	workbook.close()

def zip_files(folder_path, zip_name):
	zip_path = os.path.join(folder_path, zip_name)
	with zipfile.ZipFile(zip_path, 'w') as zipf:
		for root, _, files in os.walk(folder_path):
			for file in files:
				if file != zip_name:  # Avoid zipping the zip file itself
					file_path = os.path.join(root, file)
					arcname = os.path.relpath(file_path, folder_path)  # Relative path within the zip file
					zipf.write(file_path, arcname)

# Creating the dictionary for the Production PVPPs locations
locations_PVPPs = {"Lunca": {"lat": 46.427350, "lon": 23.905963}, "Brasov": {"lat": 45.642680, "lon": 25.588725},
					"Santimbru": {"lat":46.135244 , "lon":23.644428 }, "Bocsa": {"lat":45.377012 , "lon":21.718752}}

def render_production_forecast_Transavia(locations_PVPPs):
	st.write("Production Forecast")
	# ... (other content and functionality for production forecasting)
	# Iterating through the dictionary of PVPP locations
	for location in locations_PVPPs.keys():
		print("Getting data for {}".format(location))
		output_path = f"./Transavia/data/{location}.csv"
		lat = locations_PVPPs[location]["lat"]
		lon = locations_PVPPs[location]["lon"]
		fetch_data(lat, lon, solcast_api_key, output_path)
		# Adjusting the values to EET time
		data = pd.read_csv(f"./Transavia/data/{location}.csv")

		# Assuming 'period_end' is the column to keep fixed and all other columns are to be shifted
		columns_to_shift = data.columns.difference(['period_end'])

		# Shift the data columns by 2 intervals
		data_shifted = data[columns_to_shift].shift(2)

		# Combine the fixed 'period_end' with the shifted data columns
		data_adjusted = pd.concat([data[['period_end']], data_shifted], axis=1)

		# Optionally, handle the NaN values in the first two rows after shifting
		data_adjusted.fillna(0, inplace=True)  # Or use another method as appropriate

		# Save the adjusted DataFrame
		data_adjusted.to_csv(f"./Transavia/data/{location}.csv", index=False)
	# Creating the input_production file
	path = "./Transavia/data"
	creating_input_production_file(path)
	df = pd.read_excel("./Transavia/Production/Input_production_filled.xlsx")
	# uploaded_files = st.file_uploader("Choose a file", type=["text/csv", "xlsx"], accept_multiple_files=True)

	# if uploaded_files is not None:
	# 	for uploaded_file in uploaded_files:
	# 		if uploaded_file.type == "text/csv":
	# 			df = pd.read_csv(uploaded_file)
	# 		elif uploaded_file.type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
	# 			try:
	# 				df = pd.read_excel(uploaded_file)
	# 			except ValueError:
	# 				st.error("Expected sheet name 'Forecast_Dataset' not found in Excel file.")
	# 				continue
	# 		else:
	# 			st.error("Unsupported file format. Please upload a CSV or XLSX file.")
	# 			continue
	st.dataframe(df)
	# Submit button
	if st.button('Submit'):
		st.success('Forecast Ready', icon="✅")
		# Your code to generate the forecast
		predicting_exporting_Transavia(df)
		print("Forecast is on going...")
		# Creating the ZIP file with the Productions:
		folder_path = './Transavia/Production/Results'
		zip_name = 'Transavia_Production_Results.zip'
		zip_files(folder_path, zip_name)
		file_path = './Transavia/Production/Results/Transavia_Production_Results.zip'

		with open(file_path, "rb") as f:
			zip_data = f.read()

		# Create a download link
		b64 = base64.b64encode(zip_data).decode()
		button_html = f"""
			 <a download="Transavia_Production_Results.zip" href="data:application/zip;base64,{b64}" download>
			 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
			 </a> 
			 """
		st.markdown(button_html, unsafe_allow_html=True)

# Creating the dictionary for the Transavia locations
locations_cons = {"Lunca": {"lat": 46.427350, "lon": 23.905963}, "Brasov": {"lat": 45.642680, "lon": 25.588725},
					"Santimbru": {"lat":46.135244 , "lon":23.644428 }, "Bocsa": {"lat":45.377012 , "lon":21.718752}, "Cristian": {"lat":45.782114 , "lon":24.029499},
					"Cristuru": {"lat":46.292453 , "lon":25.031714}, "Jebel": {"lat":45.562394 , "lon":21.214496}, "Medias": {"lat":46.157283 , "lon":24.347167},
					"Miercurea": {"lat":45.890054 , "lon":23.791766}}

def render_consumption_forecast_Transavia():
	if st.button("Bring The Data"):
		st.write("Consumption Forecast")
		# ... (other content and functionality for production forecasting)
		# Iterating through the dictionary of PVPP locations
		for location in locations_cons.keys():
			print("Getting data for {}".format(location))
			output_path = f"./Transavia/data/{location}.csv"
			lat = locations_cons[location]["lat"]
			lon = locations_cons[location]["lon"]
			fetch_data(lat, lon, solcast_api_key, output_path)
		# Creating the input_production file
		path = "./Transavia/data"
		cleaning_input_files()
		creating_input_consumption_Santimbru()
		creating_input_cons_file_Brasov()
		df_santimbru = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
		df_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx")

		st.dataframe(df_santimbru)
		st.dataframe(df_brasov)
		# Creating the ZIP file with the Predictions:
		folder_path = './Transavia/Consumption/Input'
		zip_name = 'Transavia_Inputs.zip'
		zip_files(folder_path, zip_name)
		file_path = './Transavia/Consumption/Input/Transavia_Inputs.zip'

		with open(file_path, "rb") as f:
			zip_data = f.read()

		# Create a download link
		b64 = base64.b64encode(zip_data).decode()
		button_html = f"""
			 <a download="Transavia_Inputs.zip" href="data:application/zip;base64,{b64}" download>
			 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Input Files</button>
			 </a> 
			 """
		st.markdown(button_html, unsafe_allow_html=True)
	# Submit button
	if st.button('Submit'):
		st.success('Forecast Ready', icon="✅")
		# Your code to generate the forecast
		df_santimbru = pd.read_excel("./Transavia/Consumption/Input/Input.xlsx")
		df_brasov = pd.read_excel("./Transavia/Consumption/Input/Input_Brasov.xlsx")
		predicting_exporting_consumption_Santimbru(df_santimbru)
		predicting_exporting_consumption_Brasov(df_brasov)
		# Creating the ZIP file with the Predictions:
		folder_path = './Transavia/Consumption/Results/XGB'
		zip_name = 'Transavia_Consumption_Results.zip'
		zip_files(folder_path, zip_name)
		file_path = './Transavia/Consumption/Results/XGB/Transavia_Consumption_Results.zip'

		with open(file_path, "rb") as f:
			zip_data = f.read()

		# Create a download link
		b64 = base64.b64encode(zip_data).decode()
		button_html = f"""
			 <a download="Transavia_Consumption_Results.zip" href="data:application/zip;base64,{b64}" download>
			 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
			 </a> 
			 """
		st.markdown(button_html, unsafe_allow_html=True)

def render_Transavia_page():
	
	# Web App Title
	st.markdown('''
	## **The Transavia Forecast Section**

	''')

	# Allow the user to choose between Consumption and Production
	forecast_type = st.radio("Choose Forecast Type:", options=["Consumption", "Production"])

	if forecast_type == "Consumption":
		render_consumption_forecast_Transavia()
	elif forecast_type == "Production":
		render_production_forecast_Transavia(locations_PVPPs)

#================================================================================GENERAL FUNCTIONALITY==========================================================================================

#===================Creating the function for fething the data for Solina===========================
def fetching_Solina_data():
	lat = 46.073272
	lon = 23.580489
	# Fetch data from the API
	api_url = "https://api.solcast.com.au/data/forecast/radiation_and_weather?latitude={}&longitude={}&hours=168&output_parameters=air_temp,cloud_opacity,ghi&period=PT60M&format=csv&api_key={}".format(lat, lon, solcast_api_key)
	response = requests.get(api_url)
	print("Fetching data...")
	if response.status_code == 200:
		# Write the content to a CSV file
		with open("./Solina/Solcast/Alba_Iulia_raw.csv", 'wb') as file:
			file.write(response.content)
	else:
		print(response.text)  # Add this line to see the error message returned by the API
		raise Exception(f"Failed to fetch data: Status code {response.status_code}")
	# Adjusting the values to EET time
	data = pd.read_csv("./Solina/Solcast/Alba_Iulia_raw.csv")

	# Assuming 'period_end' is the column to keep fixed and all other columns are to be shifted
	columns_to_shift = data.columns.difference(['period_end'])

	# Shift the data columns by 2 intervals
	data_shifted = data[columns_to_shift].shift(2)

	# Combine the fixed 'period_end' with the shifted data columns
	data_adjusted = pd.concat([data[['period_end']], data_shifted], axis=1)

	# Optionally, handle the NaN values in the first two rows after shifting
	data_adjusted.fillna(0, inplace=True)  # Or use another method as appropriate

	# Save the adjusted DataFrame
	data_adjusted.to_csv("./Solina/Solcast/Alba_Iulia_raw.csv", index=False)


def fetching_RAAL_data():
	lat = 47.2229
	lon = 24.7244
	# Fetch data from the API
	api_url = "https://api.solcast.com.au/data/forecast/radiation_and_weather?latitude={}&longitude={}&hours=168&output_parameters=air_temp,cloud_opacity,ghi&period=PT60M&format=csv&api_key={}".format(lat, lon, solcast_api_key)
	response = requests.get(api_url)
	print("Fetching data...")
	if response.status_code == 200:
		# Write the content to a CSV file
		with open("./RAAL/Solcast/Prundu_raw.csv", 'wb') as file:
			file.write(response.content)
	else:
		print(response.text)  # Add this line to see the error message returned by the API
		raise Exception(f"Failed to fetch data: Status code {response.status_code}")
	# Adjusting the values to EET time
	data = pd.read_csv("./RAAL/Solcast/Prundu_raw.csv")

	# Assuming 'period_end' is the column to keep fixed and all other columns are to be shifted
	columns_to_shift = data.columns.difference(['period_end'])

	# Shift the data columns by 2 intervals
	data_shifted = data[columns_to_shift].shift(2)

	# Combine the fixed 'period_end' with the shifted data columns
	data_adjusted = pd.concat([data[['period_end']], data_shifted], axis=1)

	# Optionally, handle the NaN values in the first two rows after shifting
	data_adjusted.fillna(0, inplace=True)  # Or use another method as appropriate

	# Save the adjusted DataFrame
	data_adjusted.to_csv("./RAAL/Solcast/Prundu_raw.csv", index=False)

def fetching_Astro_Imperial_data():
	lat = 46.914895
	lon = 23.815583
	# Fetch data from the API
	api_url = "https://api.solcast.com.au/data/forecast/radiation_and_weather?latitude={}&longitude={}&hours=168&output_parameters=air_temp,cloud_opacity,ghi&period=PT60M&format=csv&api_key={}".format(lat, lon, solcast_api_key)
	response = requests.get(api_url)
	print("Fetching data...")
	if response.status_code == 200:
		# Write the content to a CSV file
		with open("./Astro/Solcast/Bontida_raw.csv", 'wb') as file:
			file.write(response.content)
	else:
		print(response.text)  # Add this line to see the error message returned by the API
		raise Exception(f"Failed to fetch data: Status code {response.status_code}")
	# Adjusting the values to EET time
	data = pd.read_csv("./Astro/Solcast/Bontida_raw.csv")

	# Assuming 'period_end' is the column to keep fixed and all other columns are to be shifted
	columns_to_shift = data.columns.difference(['period_end'])

	# Shift the data columns by 2 intervals
	data_shifted = data[columns_to_shift].shift(2)

	# Combine the fixed 'period_end' with the shifted data columns
	data_adjusted = pd.concat([data[['period_end']], data_shifted], axis=1)

	# Optionally, handle the NaN values in the first two rows after shifting
	data_adjusted.fillna(0, inplace=True)  # Or use another method as appropriate

	# Save the adjusted DataFrame
	data_adjusted.to_csv("./Astro/Solcast/Bontida_raw.csv", index=False)

def fetching_Astro_Imperial_data_15min():
	lat = 46.860370
	lon = 23.795201
	# Fetch data from the API
	api_url = "https://api.solcast.com.au/data/forecast/radiation_and_weather?latitude={}&longitude={}&hours=168&output_parameters=air_temp,cloud_opacity,ghi&period=PT15M&format=csv&time_zone=3&api_key={}".format(lat, lon, solcast_api_key)
	response = requests.get(api_url)
	print("Fetching data...")
	if response.status_code == 200:
		# Write the content to a CSV file
		with open("./Astro/Solcast/Jucu_15min.csv", 'wb') as file:
			file.write(response.content)
	else:
		print(response.text)  # Add this line to see the error message returned by the API
		raise Exception(f"Failed to fetch data: Status code {response.status_code}")
	# Adjusting the values to EET time
	data = pd.read_csv("./Astro/Solcast/Jucu_15min.csv")

def fetching_Astro_Imperial_data_past_15min():
	lat = 46.860370
	lon = 23.795201
	# Fetch data from the API
	api_url = "https://api.solcast.com.au/data/live/radiation_and_weather?latitude={}&longitude={}&hours=168&output_parameters=air_temp,ghi,cloud_opacity&period=PT15M&format=csv&time_zone=3&api_key={}".format(lat, lon, solcast_api_key)
	response = requests.get(api_url)
	print("Fetching data...")
	if response.status_code == 200:
		# Write the content to a CSV file
		with open("./Astro/Solcast/Jucu_15min_past.csv", 'wb') as file:
			file.write(response.content)
	else:
		print(response.text)  # Add this line to see the error message returned by the API
		raise Exception(f"Failed to fetch data: Status code {response.status_code}")
	# Adjusting the values to EET time
	data = pd.read_csv("./Astro/Solcast/Jucu_15min_past.csv")

def predicting_exporting_Astro():
	# Creating the forecast_dataset df
	data = pd.read_csv("./Astro/Solcast/Bontida_raw.csv")
	forecast_dataset = pd.read_excel("./Astro/Input_Astro.xlsx", sheet_name="Forecast_Dataset")
	# Convert 'period_end' in santimbru to datetime
	data['period_end'] = pd.to_datetime(data['period_end'], errors='coerce')
	# Extract just the date part in the desired format (as strings)
	dates = data['period_end'].dt.strftime('%Y-%m-%d')
	# Write the dates to the Input file
	forecast_dataset['Data'] = dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	forecast_dataset['Data'].fillna(method='bfill', inplace=True)
	# Completing the Interval column
	intervals = data["period_end"].dt.hour
	forecast_dataset["Interval"] = intervals
	# Replace NaNs in the 'Interval' column with 0
	forecast_dataset['Interval'].fillna(0, inplace=True)
	# Completing the Temperatura column
	forecast_dataset["Temperatura"] = data["air_temp"].values
	# Completing the GHI column
	forecast_dataset["Radiatie"] = data["ghi"].values
	# Completing the Nori column
	forecast_dataset["Nori"] = data["cloud_opacity"].values


	xgb_loaded = joblib.load("./Astro/rs_xgb_Astro_prod.pkl")

	forecast_dataset["Month"] = pd.to_datetime(forecast_dataset.Data).dt.month
	dataset = forecast_dataset.copy()
	forecast_dataset = forecast_dataset.drop("Data", axis=1)

	preds = xgb_loaded.predict(forecast_dataset.values)
	
	# Rounding each value in the list to the third decimal
	rounded_values = [round(value, 3) for value in preds]
	
	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Astro/Results_Production_Astro_xgb.xlsx")
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in rounded_values:
			worksheet.write(row, col + 2, value, decimal_format)
			row +=1
	row = 1
	for Data, Interval in zip(dataset.Data, dataset.Interval):
			worksheet.write(row, col + 0, Data, date_format)
			worksheet.write(row, col + 1, Interval)
			row +=1

	workbook.close()
	# Formatting the Results file
	# Step 1: Open the Excel file
	file_path = "./Astro/Results_Production_Astro_xgb.xlsx"
	workbook = load_workbook(filename=file_path)
	worksheet = workbook['Production_Predictions']  # Adjust the sheet name as necessary

	# Step 2: Directly round the values in column C and write them back
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value is not None:  # Check if the cell is not empty
			# Round the value to 3 decimal places and write it back to column C
			worksheet.cell(row, 3).value = round(original_value, 3)
		
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value < 0.01:  # Check if the value is less than 0.01
			# Residual values are rounded to 0.000
			worksheet.cell(row, 3).value = 0
	# Save the workbook with the rounded values
	workbook.save(filename=file_path)
	workbook.close()
	# Open the existing workbook
	# Load the Excel file into a DataFrame
	df = pd.read_excel(file_path)
	
	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])
	
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Interval"].astype(str)
	df.to_excel(file_path, index=False)
	return dataset

def predicting_exporting_Astro_15min():
	# Creating the forecast_dataset df
	df= pd.read_csv('./Astro/Solcast/Jucu_15min.csv')
	# Convert the 'period_end' column to datetime, handling errors
	df['period_end'] = pd.to_datetime(df['period_end'], errors='coerce', format='%Y-%m-%dT%H:%M:%SZ')

	# Drop any rows with NaT in 'period_end'
	df.dropna(subset=['period_end'], inplace=True)

	# Shift the 'period_end' column by 2 hours
	df['period_end'] = df['period_end'] + pd.Timedelta(hours=3)

	# Creating the Interval column
	df['Interval'] = df.period_end.dt.hour * 4 + df.period_end.dt.minute // 15 + 1

	df.rename(columns={'period_end': 'Data', 'ghi': 'Radiatie', "air_temp": "Temperatura", "cloud_opacity": "Nori"}, inplace=True)

	df = df[["Data", "Interval", "Temperatura", "Nori", "Radiatie"]]

	xgb_loaded = joblib.load("./Astro/rs_xgb_Astro_prod_15min.pkl")

	df["Month"] = df.Data.dt.month
	dataset = df.copy()
	forecast_dataset = dataset[["Interval", "Temperatura", "Nori", "Radiatie", "Month"]]

	preds = xgb_loaded.predict(forecast_dataset.values)
	
	# Rounding each value in the list to the third decimal
	rounded_values = [round(value, 3) for value in preds]
	
	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Astro/Results_Production_Astro_xgb_15min.xlsx")
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in rounded_values:
			worksheet.write(row, col + 2, value, decimal_format)
			row +=1
	row = 1
	for Data, Interval in zip(dataset.Data, dataset.Interval):
			worksheet.write(row, col + 0, Data, date_format)
			worksheet.write(row, col + 1, Interval)
			row +=1

	workbook.close()
	# Formatting the Results file
	# Step 1: Open the Excel file
	file_path = "./Astro/Results_Production_Astro_xgb_15min_0624.xlsx"
	workbook = load_workbook(filename=file_path)
	worksheet = workbook['Production_Predictions']  # Adjust the sheet name as necessary

	# Step 2: Directly round the values in column C and write them back
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value is not None:  # Check if the cell is not empty
			# Round the value to 3 decimal places and write it back to column C
			worksheet.cell(row, 3).value = round(original_value, 3)
		
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value < 0.01:  # Check if the value is less than 0.01
			# Residual values are rounded to 0.000
			worksheet.cell(row, 3).value = 0
	# Save the workbook with the rounded values
	workbook.save(filename=file_path)
	workbook.close()
	# Open the existing workbook
	# Load the Excel file into a DataFrame
	df = pd.read_excel(file_path)
	
	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])
	
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Interval"].astype(str)
	df.to_excel(file_path, index=False)
	return dataset

# Function to handle file upload of the real-time consumption data
def upload_file():
	uploaded_file = st.file_uploader("Upload Real-Time Production Data File", type=['csv', 'xlsx'])
	if uploaded_file is not None:
		if uploaded_file.type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':  # Excel file
			data = pd.read_excel(uploaded_file)
		elif uploaded_file.type == 'text/csv':  # CSV file
			data = pd.read_csv(uploaded_file)
		
		# Process the data (e.g., display first few rows)
		if data is not None:
			st.write("Uploaded file:")
			st.write(data.head())
		else:
			st.write("Unsupported file format. Please upload a CSV or Excel file.")

def predicting_exporting_Astro_Intraday_15min(real_time_data):
	# Creating the forecast_dataset df
	weather_data_future = pd.read_csv('./Astro/Solcast/Jucu_15min.csv')
	# Convert the 'period_end' column to datetime, handling errors
	weather_data_future['period_end'] = pd.to_datetime(weather_data_future['period_end'], errors='coerce', format='%Y-%m-%dT%H:%M:%SZ')

	# Drop any rows with NaT in 'period_end'
	weather_data_future.dropna(subset=['period_end'], inplace=True)

	# Shift the 'period_end' column by 3 hours
	weather_data_future['period_end'] = weather_data_future['period_end'] + pd.Timedelta(hours=3)

	# Creating the Interval column
	weather_data_future['Interval'] = weather_data_future.period_end.dt.hour * 4 + weather_data_future.period_end.dt.minute // 15 + 1

	weather_data_future.rename(columns={'period_end': 'Timestamp', 'ghi': 'Radiatie', "air_temp": "Temperatura", "cloud_opacity": "Nori"}, inplace=True)

	weather_data_future["Month"] = weather_data_future.Timestamp.dt.month

	weather_data_future = weather_data_future[["Timestamp", "Interval", "Temperatura", "Nori", "Radiatie", "Month"]]

	# Data Engineering for the past weather data
	weather_data_past = pd.read_csv('./Astro/Solcast/Jucu_15min_past.csv')
	# Convert the 'period_end' column to datetime, handling errors
	weather_data_past['period_end'] = pd.to_datetime(weather_data_past['period_end'], errors='coerce', format='%Y-%m-%dT%H:%M:%SZ')

	# Drop any rows with NaT in 'period_end'
	weather_data_past.dropna(subset=['period_end'], inplace=True)

	# Shift the 'period_end' column by 2 hours
	weather_data_past['period_end'] = weather_data_past['period_end'] + pd.Timedelta(hours=3)

	# Rename 'Data' column to 'Timestamp' for consistency
	weather_data_past.rename(columns={'period_end': 'Timestamp', 'ghi': 'Radiatie', "air_temp": "Temperatura", "cloud_opacity": "Nori"}, inplace=True)

	# Creating the Interval column
	weather_data_past['Interval'] = weather_data_past.Timestamp.dt.hour * 4 + weather_data_past.Timestamp.dt.minute // 15 + 1

	weather_data_past['Month'] = weather_data_past.Timestamp.dt.month

	# Reorder the columns
	weather_data_past = weather_data_past[['Timestamp', "Interval", "Temperatura", "Nori", "Radiatie", "Month"]]

	# Concatenating the two weather datasets
	# Ensure 'Timestamp' column in weather_data_future is in datetime format
	weather_data_future['Timestamp'] = pd.to_datetime(weather_data_future['Timestamp'])

	# Concatenate dataframes
	weather_data = pd.concat([weather_data_future, weather_data_past], ignore_index=True)

	# Step 2: Sort the DataFrame by 'Timestamp' in ascending order
	weather_data.sort_values('Timestamp', inplace=True)

	# Step 3: Remove duplicates - assuming you want to keep the first occurrence of each 'Timestamp'
	weather_data.drop_duplicates(subset='Timestamp', keep='first', inplace=True)
	st.text("The concatenated weather data for Astro")
	st.dataframe(weather_data)

	# Data Engineering the Real-Time Production data
	# Load your data
	real_time_data = real_time_data.copy()

	# Replace '--' with 0 in the 'Power' column
	real_time_data['Power'] = real_time_data['Power'].replace('--', 0)

	# Convert the 'Power' column to numeric (will convert invalid parsing to NaN)
	real_time_data['Power'] = pd.to_numeric(real_time_data['Power'], errors='coerce').fillna(0)

	# Assuming the 'Power' data needs to be in megawatts and currently in watts
	real_time_data['Power_MW'] = real_time_data['Power'] / 1000000  # Convert W to MW

	# Ensure the 'Timestamp' column is in datetime format
	real_time_data['Timestamp'] = pd.to_datetime(real_time_data['Timestamp'])

	# Shift the 'Timestamp' column by 3 hours forward
	real_time_data['Timestamp'] = real_time_data['Timestamp'] + pd.Timedelta(hours=3)

	# Calculate the average of current and next power readings
	real_time_data['Next_Power_MW'] = real_time_data['Power_MW'].shift(1)  # Shift upwards to get the next reading in the column

	# Calculate the average power
	real_time_data['Average_Power_MW'] = (real_time_data['Power_MW'] + real_time_data['Next_Power_MW']) / 2

	# Calculate the energy for each interval
	real_time_data['Energy_MWh'] = real_time_data['Average_Power_MW'] * 0.25

	# Convert the 'Timestamp' column to timezone-naive datetime objects
	real_time_data['Timestamp'] = pd.to_datetime(real_time_data['Timestamp']).dt.tz_localize(None)
	st.text("The Real-Time Production")
	st.dataframe(real_time_data)
	# Merge datasets on the 'Timestamp' column
	forecast_dataset = pd.merge(real_time_data, weather_data, on='Timestamp', how='inner')
	st.text("This is the Forecast Dataset")
	st.dataframe(forecast_dataset)

	# Creating rolling and lag features
	forecast_dataset['Temp_Rolling_Avg'] = forecast_dataset['Temperatura'].rolling(window=4).mean()  # 1-hour rolling average
	forecast_dataset['Radiation_Rolling_Avg'] = forecast_dataset['Radiatie'].rolling(window=4).mean()  # 1-hour rolling average

	# Renaming columns
	forecast_dataset.rename(columns={'Energy_MWh': 'Productie'}, inplace=True)

	# Creating lag features for the last four 15-minute intervals
	for i in range(1, 5):
		forecast_dataset[f'Productie_Lag_{i}'] = forecast_dataset['Productie'].shift(i)
		forecast_dataset[f'Radiatie_Lag_{i}'] = forecast_dataset['Radiatie'].shift(i)

	dataset = forecast_dataset.copy()

	forecast_dataset = forecast_dataset[['Interval', 'Temperatura', 'Nori', 'Radiatie', 'Month',
	   'Temp_Rolling_Avg', 'Radiation_Rolling_Avg', 'Productie_Lag_1',
	   'Radiatie_Lag_1', 'Productie_Lag_2', 'Radiatie_Lag_2',
	   'Productie_Lag_3', 'Radiatie_Lag_3', 'Productie_Lag_4',
	   'Radiatie_Lag_4']]

	forecast_dataset.dropna(inplace=True)

	st.dataframe(forecast_dataset)

	xgb_loaded = joblib.load("./Astro/rs_xgb_Astro_intraday_prod_15min.pkl")

	preds = xgb_loaded.predict(forecast_dataset.values)
	
	# Rounding each value in the list to the third decimal
	rounded_values = [round(value, 3) for value in preds]
	
	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Astro/Results_Production_Astro_xgb_intraday_15min.xlsx")
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in rounded_values:
			worksheet.write(row, col + 2, value, decimal_format)
			row +=1
	row = 1
	for data, interval in zip(dataset.Timestamp, dataset.Interval):
		if interval <= 91:
			worksheet.write(row, col + 0, data, date_format)
			worksheet.write(row, col + 1, interval + 5)
			row +=1

	workbook.close()
	# Formatting the Results file
	# Step 1: Open the Excel file
	file_path = "./Astro/Results_Production_Astro_xgb_intraday_15min.xlsx"
	workbook = load_workbook(filename=file_path)
	worksheet = workbook['Production_Predictions']  # Adjust the sheet name as necessary

	# Step 2: Directly round the values in column C and write them back
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value is not None:  # Check if the cell is not empty
			# Round the value to 3 decimal places and write it back to column C
			worksheet.cell(row, 3).value = round(original_value, 3)
		
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value < 0.01:  # Check if the value is less than 0.01
			# Residual values are rounded to 0.000
			worksheet.cell(row, 3).value = 0
	# Save the workbook with the rounded values
	workbook.save(filename=file_path)
	workbook.close()
	# Open the existing workbook
	# Load the Excel file into a DataFrame
	df = pd.read_excel(file_path)
	
	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])
	
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Interval"].astype(str)
	df.to_excel(file_path, index=False)
	return dataset

# Function to make dynamic predictions using the strongest feature
def make_dynamic_predictions(model, real_time_data, start_interval, end_interval):
	predictions = []
	available_data = real_time_data.copy()

	# Use the last real production value to start predictions
	for interval in range(start_interval, end_interval + 1):
		available_data = create_lag_features(available_data)
		
		# Check if there's enough data to proceed
		if len(available_data) == 0:
			print(f"Insufficient data at interval {interval}")
			break
		
		# Debugging statements to track the state of available_data
		print(f"Interval: {interval}")
		print(f"Available data size: {len(available_data)}")
		print(available_data.tail())
		
		# Get the latest features for the prediction interval
		latest_data = available_data.iloc[-1]
		features = latest_data[['Interval', 'Temperatura', 'Nori', 'Radiatie', 'Month', 'Productie_lag1', 'Productie_lag2', 'Productie_lag3', 'Productie_rolling_mean', 'Productie_rolling_std']].values.reshape(1, -1)

		# Make prediction
		prediction = model.predict(features)[0]
		prediction = max(prediction, 0)  # Replace negative predictions with zero
		predictions.append((interval, prediction))

		# Update the available_data with the new prediction
		new_data = pd.DataFrame({
			'Productie': [prediction],
			'Interval': [interval],
			'Temperatura': [latest_data['Temperatura']],
			'Nori': [latest_data['Nori']],
			'Radiatie': [latest_data['Radiatie']],
			'Month': [latest_data['Month']]
		})
		available_data = pd.concat([available_data, new_data], ignore_index=True)

	return predictions

# Function to create lag features dynamically
def create_lag_features(data):
	data['Productie_lag1'] = data['Productie'].shift(1)
	return data.dropna()

def predicting_rest_of_day(interval, data, model):
	# Creating the lag features
	create_lag_features(data)
	features = data[['Interval', 'Temperatura', 'Nori', 'Radiatie', 'Month', 'Productie_lag1']]
	features.dropna(inplace=True)
	# Now we need to forecast the Production for the last real interval which will become the first Production_lag1 for the next one
	interval_features = features[features['Interval'] == interval]
	preds_interval = model.predict(interval_features.values)
	# Now, we need to add this prediction value to the forecast dataset instead of the last interval with Productie > 0
	# Add the prediction to the dataset
	data.loc[data['Interval'] == interval, 'Productie'] = preds_interval
	return data


def predicting_exporting_Imperial_Intraday_15min(real_time_data):
	# Creating the forecast_dataset df
	weather_data_future = pd.read_csv('./Astro/Solcast/Jucu_15min.csv')
	# Convert the 'period_end' column to datetime, handling errors
	weather_data_future['period_end'] = pd.to_datetime(weather_data_future['period_end'], errors='coerce', format='%Y-%m-%dT%H:%M:%SZ')

	# Drop any rows with NaT in 'period_end'
	weather_data_future.dropna(subset=['period_end'], inplace=True)

	# Shift the 'period_end' column by 3 hours
	weather_data_future['period_end'] = weather_data_future['period_end'] + pd.Timedelta(hours=3)

	# Creating the Interval column
	weather_data_future['Interval'] = weather_data_future.period_end.dt.hour * 4 + weather_data_future.period_end.dt.minute // 15 + 1

	weather_data_future.rename(columns={'period_end': 'Timestamp', 'ghi': 'Radiatie', "air_temp": "Temperatura", "cloud_opacity": "Nori"}, inplace=True)

	weather_data_future["Month"] = weather_data_future.Timestamp.dt.month

	weather_data_future = weather_data_future[["Timestamp", "Interval", "Temperatura", "Nori", "Radiatie", "Month"]]

	# Data Engineering for the past weather data
	weather_data_past = pd.read_csv('./Astro/Solcast/Jucu_15min_past.csv')
	# Convert the 'period_end' column to datetime, handling errors
	weather_data_past['period_end'] = pd.to_datetime(weather_data_past['period_end'], errors='coerce', format='%Y-%m-%dT%H:%M:%SZ')

	# Drop any rows with NaT in 'period_end'
	weather_data_past.dropna(subset=['period_end'], inplace=True)

	# Shift the 'period_end' column by 2 hours
	weather_data_past['period_end'] = weather_data_past['period_end'] + pd.Timedelta(hours=3)

	# Rename 'Data' column to 'Timestamp' for consistency
	weather_data_past.rename(columns={'period_end': 'Timestamp', 'ghi': 'Radiatie', "air_temp": "Temperatura", "cloud_opacity": "Nori"}, inplace=True)

	# Creating the Interval column
	weather_data_past['Interval'] = weather_data_past.Timestamp.dt.hour * 4 + weather_data_past.Timestamp.dt.minute // 15 + 1

	weather_data_past['Month'] = weather_data_past.Timestamp.dt.month

	# Reorder the columns
	weather_data_past = weather_data_past[['Timestamp', "Interval", "Temperatura", "Nori", "Radiatie", "Month"]]

	# Concatenating the two weather datasets
	# Ensure 'Timestamp' column in weather_data_future is in datetime format
	weather_data_future['Timestamp'] = pd.to_datetime(weather_data_future['Timestamp'])

	# Concatenate dataframes
	weather_data = pd.concat([weather_data_future, weather_data_past], ignore_index=True)

	# Step 2: Sort the DataFrame by 'Timestamp' in ascending order
	weather_data.sort_values('Timestamp', inplace=True)

	# Step 3: Remove duplicates - assuming you want to keep the first occurrence of each 'Timestamp'
	weather_data.drop_duplicates(subset='Timestamp', keep='first', inplace=True)

	st.dataframe(weather_data)

	st.dataframe(real_time_data)

	# Convert the 'Timestamp' column to timezone-naive datetime objects
	real_time_data['Timestamp'] = pd.to_datetime(real_time_data['Timestamp']).dt.tz_localize(None)

	# Merge datasets on the 'Timestamp' column
	forecast_dataset = pd.merge(real_time_data, weather_data, on='Timestamp', how='inner')

	forecast_dataset.rename(columns={'Energy_MWh': 'Productie'}, inplace=True)

	st.dataframe(forecast_dataset)

	dataset = forecast_dataset.copy()

	# Find the last interval with real production data 
	last_real_interval = forecast_dataset[forecast_dataset['Productie'] > 0]['Interval'].max()
	model = joblib.load("./Imperial/rs_xgb_Imperial_prod_intraday_1lag_15min.pkl")

	for interval in range(last_real_interval, 97):
		print(forecast_dataset)
		forecast_dataset = predicting_rest_of_day(interval, forecast_dataset, model)
		print(f"Predicted production for interval {interval}: {forecast_dataset.loc[forecast_dataset['Interval'] == interval, 'Productie'].values[0]}")

	st.text("This is the forecast dataset after predictions:")

	forecast_dataset.loc[forecast_dataset['Radiatie'] == 0, 'Productie'] = 0
	st.dataframe(forecast_dataset)

	preds = forecast_dataset.Productie.values
	preds = np.nan_to_num(preds, nan=0)
	st.write(len(preds))
	# Rounding each value in the list to the third decimal
	rounded_values = [round(value, 3) for value in preds]
	
	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Imperial/Results_Production_Imperial_xgb_intraday_15min.xlsx")
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in rounded_values:
			worksheet.write(row, col + 2, value, decimal_format)
			row +=1
	row = 1
	for data, interval in zip(dataset.Timestamp, dataset.Interval):
		worksheet.write(row, col + 0, data, date_format)
		worksheet.write(row, col + 1, interval)
		row +=1

	workbook.close()
	# Formatting the Results file
	# Step 1: Open the Excel file
	file_path = "./Imperial/Results_Production_Imperial_xgb_intraday_15min.xlsx"
	workbook = load_workbook(filename=file_path)
	worksheet = workbook['Production_Predictions']  # Adjust the sheet name as necessary

	# Step 2: Directly round the values in column C and write them back
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value is not None:  # Check if the cell is not empty
			# Round the value to 3 decimal places and write it back to column C
			worksheet.cell(row, 3).value = round(original_value, 3)
		
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value < 0.01:  # Check if the value is less than 0.01
			# Residual values are rounded to 0.000
			worksheet.cell(row, 3).value = 0
	# Save the workbook with the rounded values
	workbook.save(filename=file_path)
	workbook.close()
	# Open the existing workbook
	# Load the Excel file into a DataFrame
	df = pd.read_excel(file_path)
	
	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])
	
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Interval"].astype(str)
	df.to_excel(file_path, index=False)
	return dataset

def predicting_exporting_Astro_Intraday_15min(real_time_data):
	# Creating the forecast_dataset df
	weather_data_future = pd.read_csv('./Astro/Solcast/Jucu_15min.csv')
	# Convert the 'period_end' column to datetime, handling errors
	weather_data_future['period_end'] = pd.to_datetime(weather_data_future['period_end'], errors='coerce', format='%Y-%m-%dT%H:%M:%SZ')

	# Drop any rows with NaT in 'period_end'
	weather_data_future.dropna(subset=['period_end'], inplace=True)

	# Shift the 'period_end' column by 3 hours
	weather_data_future['period_end'] = weather_data_future['period_end'] + pd.Timedelta(hours=3)

	# Creating the Interval column
	weather_data_future['Interval'] = weather_data_future.period_end.dt.hour * 4 + weather_data_future.period_end.dt.minute // 15 + 1

	weather_data_future.rename(columns={'period_end': 'Timestamp', 'ghi': 'Radiatie', "air_temp": "Temperatura", "cloud_opacity": "Nori"}, inplace=True)

	weather_data_future["Month"] = weather_data_future.Timestamp.dt.month

	weather_data_future = weather_data_future[["Timestamp", "Interval", "Temperatura", "Nori", "Radiatie", "Month"]]

	# Data Engineering for the past weather data
	weather_data_past = pd.read_csv('./Astro/Solcast/Jucu_15min_past.csv')
	# Convert the 'period_end' column to datetime, handling errors
	weather_data_past['period_end'] = pd.to_datetime(weather_data_past['period_end'], errors='coerce', format='%Y-%m-%dT%H:%M:%SZ')

	# Drop any rows with NaT in 'period_end'
	weather_data_past.dropna(subset=['period_end'], inplace=True)

	# Shift the 'period_end' column by 2 hours
	weather_data_past['period_end'] = weather_data_past['period_end'] + pd.Timedelta(hours=3)

	# Rename 'Data' column to 'Timestamp' for consistency
	weather_data_past.rename(columns={'period_end': 'Timestamp', 'ghi': 'Radiatie', "air_temp": "Temperatura", "cloud_opacity": "Nori"}, inplace=True)

	# Creating the Interval column
	weather_data_past['Interval'] = weather_data_past.Timestamp.dt.hour * 4 + weather_data_past.Timestamp.dt.minute // 15 + 1

	weather_data_past['Month'] = weather_data_past.Timestamp.dt.month

	# Reorder the columns
	weather_data_past = weather_data_past[['Timestamp', "Interval", "Temperatura", "Nori", "Radiatie", "Month"]]

	# Concatenating the two weather datasets
	# Ensure 'Timestamp' column in weather_data_future is in datetime format
	weather_data_future['Timestamp'] = pd.to_datetime(weather_data_future['Timestamp'])

	# Concatenate dataframes
	weather_data = pd.concat([weather_data_future, weather_data_past], ignore_index=True)

	# Step 2: Sort the DataFrame by 'Timestamp' in ascending order
	weather_data.sort_values('Timestamp', inplace=True)

	# Step 3: Remove duplicates - assuming you want to keep the first occurrence of each 'Timestamp'
	weather_data.drop_duplicates(subset='Timestamp', keep='first', inplace=True)

	st.dataframe(weather_data)

	st.dataframe(real_time_data)

	# Convert the 'Timestamp' column to timezone-naive datetime objects
	real_time_data['Timestamp'] = pd.to_datetime(real_time_data['Timestamp']).dt.tz_localize(None)

	# Merge datasets on the 'Timestamp' column
	forecast_dataset = pd.merge(real_time_data, weather_data, on='Timestamp', how='inner')

	forecast_dataset.rename(columns={'Energy_MWh': 'Productie'}, inplace=True)

	st.dataframe(forecast_dataset)

	dataset = forecast_dataset.copy()

	# Find the last interval with real production data 
	last_real_interval = forecast_dataset[forecast_dataset['Productie'] > 0]['Interval'].max()
	model = joblib.load("./Astro/rs_xgb_Astro_prod_intraday_1lag_15min.pkl")

	for interval in range(last_real_interval, 97):
		print(forecast_dataset)
		forecast_dataset = predicting_rest_of_day(interval, forecast_dataset, model)
		print(f"Predicted production for interval {interval}: {forecast_dataset.loc[forecast_dataset['Interval'] == interval, 'Productie'].values[0]}")

	
	forecast_dataset.loc[forecast_dataset['Radiatie'] == 0, 'Productie'] = 0
	st.text("This is the final dataframe:")
	st.dataframe(forecast_dataset)
	preds = forecast_dataset.Productie.values
	preds = np.nan_to_num(preds, nan=0)
	st.write(preds)
	# Rounding each value in the list to the third decimal
	rounded_values = [round(value, 3) for value in preds]
	
	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Astro/Results_Production_Astro_xgb_intraday_15min.xlsx")
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in rounded_values:
			worksheet.write(row, col + 2, value, decimal_format)
			row +=1
	row = 1
	for data, interval in zip(dataset.Timestamp, dataset.Interval):
		worksheet.write(row, col + 0, data, date_format)
		worksheet.write(row, col + 1, interval)
		row +=1

	workbook.close()
	# Formatting the Results file
	# Step 1: Open the Excel file
	file_path = "./Astro/Results_Production_Astro_xgb_intraday_15min.xlsx"
	workbook = load_workbook(filename=file_path)
	worksheet = workbook['Production_Predictions']  # Adjust the sheet name as necessary

	# Step 2: Directly round the values in column C and write them back
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value is not None:  # Check if the cell is not empty
			# Round the value to 3 decimal places and write it back to column C
			worksheet.cell(row, 3).value = round(original_value, 3)
		
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value < 0.01:  # Check if the value is less than 0.01
			# Residual values are rounded to 0.000
			worksheet.cell(row, 3).value = 0
	# Save the workbook with the rounded values
	workbook.save(filename=file_path)
	workbook.close()
	# Open the existing workbook
	# Load the Excel file into a DataFrame
	df = pd.read_excel(file_path)
	
	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])
	
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Interval"].astype(str)
	df.to_excel(file_path, index=False)
	return dataset

def creating_prediction_dataset_Astro():
	# Creating the forecast_dataset df
	data = pd.read_csv("./Astro/Solcast/Bontida_raw.csv")
	forecast_dataset = pd.read_excel("./Astro/Input_Astro.xlsx", sheet_name="Forecast_Dataset")
	# Convert 'period_end' in santimbru to datetime
	data['period_end'] = pd.to_datetime(data['period_end'], errors='coerce')
	# Extract just the date part in the desired format (as strings)
	dates = data['period_end'].dt.strftime('%Y-%m-%d')
	# Write the dates to the Input file
	forecast_dataset['Data'] = dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	forecast_dataset['Data'].fillna(method='bfill', inplace=True)
	# Completing the Interval column
	intervals = data["period_end"].dt.hour
	forecast_dataset["Interval"] = intervals
	# Replace NaNs in the 'Interval' column with 0
	forecast_dataset['Interval'].fillna(0, inplace=True)
	# Completing the Temperatura column
	forecast_dataset["Temperatura"] = data["air_temp"].values
	# Completing the GHI column
	forecast_dataset["Radiatie"] = data["ghi"].values
	# Completing the Nori column
	forecast_dataset["Nori"] = data["cloud_opacity"].values
	forecast_dataset['Data'] = pd.to_datetime(forecast_dataset['Data'])
	df_predictions = pd.read_excel("./Astro/Results_Production_Astro_xgb.xlsx")
	df_final = pd.merge(forecast_dataset, df_predictions, on=['Data', 'Interval'], how='left')
	df_final.to_excel("./Astro/predictions_dataset.xlsx", index=False)

def predicting_exporting_Imperial():
	# Creating the forecast_dataset df
	data = pd.read_csv("./Astro/Solcast/Bontida_raw.csv")
	forecast_dataset = pd.read_excel("./Imperial/Input_Imperial.xlsx", sheet_name="Forecast_Dataset")
	# Convert 'period_end' in santimbru to datetime
	data['period_end'] = pd.to_datetime(data['period_end'], errors='coerce')
	# Extract just the date part in the desired format (as strings)
	dates = data['period_end'].dt.strftime('%Y-%m-%d')
	# Write the dates to the Input file
	forecast_dataset['Data'] = dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	forecast_dataset['Data'].fillna(method='bfill', inplace=True)
	# Completing the Interval column
	intervals = data["period_end"].dt.hour
	forecast_dataset["Interval"] = intervals
	# Replace NaNs in the 'Interval' column with 0
	forecast_dataset['Interval'].fillna(0, inplace=True)
	# Completing the Temperatura column
	forecast_dataset["Temperatura"] = data["air_temp"].values
	# Completing the GHI column
	forecast_dataset["Radiatie"] = data["ghi"].values
	# Completing the Nori column
	forecast_dataset["Nori"] = data["cloud_opacity"].values


	xgb_loaded = joblib.load("./Imperial/rs_xgb_Imperial_prod_clean_data.pkl")

	forecast_dataset["Month"] = pd.to_datetime(forecast_dataset.Data).dt.month
	dataset = forecast_dataset.copy()
	forecast_dataset = forecast_dataset.drop("Data", axis=1)

	preds = xgb_loaded.predict(forecast_dataset.values)
	
	# Rounding each value in the list to the third decimal
	rounded_values = [round(value, 3) for value in preds]
	
	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Imperial/Results_Production_Imperial_xgb.xlsx")
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in rounded_values:
			worksheet.write(row, col + 2, value, decimal_format)
			row +=1
	row = 1
	for Data, Interval in zip(dataset.Data, dataset.Interval):
			worksheet.write(row, col + 0, Data, date_format)
			worksheet.write(row, col + 1, Interval)
			row +=1

	workbook.close()
	# Formatting the Results file
	# Step 1: Open the Excel file
	file_path = "./Imperial/Results_Production_Imperial_xgb.xlsx"
	workbook = load_workbook(filename=file_path)
	worksheet = workbook['Production_Predictions']  # Adjust the sheet name as necessary

	# Step 2: Directly round the values in column C and write them back
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value is not None:  # Check if the cell is not empty
			# Round the value to 3 decimal places and write it back to column C
			worksheet.cell(row, 3).value = round(original_value, 3)
		
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value < 0.01:  # Check if the value is less than 0.01
			# Residual values are rounded to 0.000
			worksheet.cell(row, 3).value = 0
	# Save the workbook with the rounded values
	workbook.save(filename=file_path)
	workbook.close()

	# Open the existing workbook
	# Load the Excel file into a DataFrame
	df = pd.read_excel(file_path)
	
	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])
	
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Interval"].astype(str)
	df.to_excel(file_path, index=False)
	return dataset

def predicting_exporting_Imperial_15min():
	# Creating the forecast_dataset df
	df= pd.read_csv('./Astro/Solcast/Jucu_15min.csv')
	# Convert the 'period_end' column to datetime, handling errors
	df['period_end'] = pd.to_datetime(df['period_end'], errors='coerce', format='%Y-%m-%dT%H:%M:%SZ')

	# Drop any rows with NaT in 'period_end'
	df.dropna(subset=['period_end'], inplace=True)

	# Shift the 'period_end' column by 2 hours
	df['period_end'] = df['period_end'] + pd.Timedelta(hours=3)

	# Creating the Interval column
	df['Interval'] = df.period_end.dt.hour * 4 + df.period_end.dt.minute // 15 + 1

	df.rename(columns={'period_end': 'Data', 'ghi': 'Radiatie', "air_temp": "Temperatura", "cloud_opacity": "Nori"}, inplace=True)

	df = df[["Data", "Interval", "Temperatura", "Nori", "Radiatie"]]

	xgb_loaded = joblib.load("./Astro/rs_xgb_Astro_prod_15min.pkl")

	df["Month"] = df.Data.dt.month
	dataset = df.copy()
	forecast_dataset = dataset[["Interval", "Temperatura", "Nori", "Radiatie", "Month"]]

	xgb_loaded = joblib.load("./Imperial/rs_xgb_Imperial_prod_15min.pkl")

	preds = xgb_loaded.predict(forecast_dataset.values)
	
	# Rounding each value in the list to the third decimal
	rounded_values = [round(value, 3) for value in preds]
	
	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Imperial/Results_Production_Imperial_xgb_15min.xlsx")
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in rounded_values:
			worksheet.write(row, col + 2, value, decimal_format)
			row +=1
	row = 1
	for Data, Interval in zip(dataset.Data, dataset.Interval):
			worksheet.write(row, col + 0, Data, date_format)
			worksheet.write(row, col + 1, Interval)
			row +=1

	workbook.close()
	# Formatting the Results file
	# Step 1: Open the Excel file
	file_path = "./Imperial/Results_Production_Imperial_xgb_15min.xlsx"
	workbook = load_workbook(filename=file_path)
	worksheet = workbook['Production_Predictions']  # Adjust the sheet name as necessary

	# Step 2: Directly round the values in column C and write them back
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value is not None:  # Check if the cell is not empty
			# Round the value to 3 decimal places and write it back to column C
			worksheet.cell(row, 3).value = round(original_value, 3)
		
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value < 0.01:  # Check if the value is less than 0.01
			# Residual values are rounded to 0.000
			worksheet.cell(row, 3).value = 0
	# Save the workbook with the rounded values
	workbook.save(filename=file_path)
	workbook.close()
	# Open the existing workbook
	# Load the Excel file into a DataFrame
	df = pd.read_excel(file_path)
	
	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])
	
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Interval"].astype(str)
	df.to_excel(file_path, index=False)
	return dataset

def predicting_exporting_Solina():
	# Creating the forecast_dataset df
	data = pd.read_csv("./Solina/Solcast/Alba_Iulia_raw.csv")
	forecast_dataset = pd.read_excel("./Solina/Production/Input_Solina.xlsx", sheet_name="Forecast_Dataset")
	# Convert 'period_end' in santimbru to datetime
	data['period_end'] = pd.to_datetime(data['period_end'], errors='coerce')
	# Extract just the date part in the desired format (as strings)
	dates = data['period_end'].dt.strftime('%Y-%m-%d')
	# Write the dates to the Input file
	forecast_dataset['Data'] = dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	forecast_dataset['Data'].fillna(method='bfill', inplace=True)
	# Completing the Interval column
	intervals = data["period_end"].dt.hour
	forecast_dataset["Interval"] = intervals
	# Replace NaNs in the 'Interval' column with 0
	forecast_dataset['Interval'].fillna(0, inplace=True)
	# Completing the Temperatura column
	forecast_dataset["Temperatura"] = data["air_temp"].values
	# Completing the GHI column
	forecast_dataset["Radiatie"] = data["ghi"].values
	# Completing the Nori column
	forecast_dataset["Nori"] = data["cloud_opacity"].values


	xgb_loaded = joblib.load("./Solina/Production/rs_xgb_Solina_prod.pkl")

	forecast_dataset["Month"] = pd.to_datetime(forecast_dataset.Data).dt.month
	dataset = forecast_dataset.copy()
	forecast_dataset = forecast_dataset.drop("Data", axis=1)

	preds = xgb_loaded.predict(forecast_dataset.values)
	
	# Rounding each value in the list to the third decimal
	rounded_values = [round(value, 3) for value in preds]
	
	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Solina/Production/Results_Production_xgb.xlsx")
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in rounded_values:
			worksheet.write(row, col + 2, value, decimal_format)
			row +=1
	row = 1
	for Data, Interval in zip(dataset.Data, dataset.Interval):
			worksheet.write(row, col + 0, Data, date_format)
			worksheet.write(row, col + 1, Interval)
			row +=1

	workbook.close()
	# Formatting the Results file
	# Step 1: Open the Excel file
	file_path = "./Solina/Production/Results_Production_xgb.xlsx"
	workbook = load_workbook(filename=file_path)
	worksheet = workbook['Production_Predictions']  # Adjust the sheet name as necessary

	# Step 2: Directly round the values in column C and write them back
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value is not None:  # Check if the cell is not empty
			# Round the value to 3 decimal places and write it back to column C
			worksheet.cell(row, 3).value = round(original_value, 3)
		
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value < 0.01:  # Check if the value is less than 0.01
			# Residual values are rounded to 0.000
			worksheet.cell(row, 3).value = 0
	# Save the workbook with the rounded values
	workbook.save(filename=file_path)
	workbook.close()
	# Open the existing workbook
	# Load the Excel file into a DataFrame
	df = pd.read_excel(file_path)
	
	# Ensure the 'Data' column is in datetime format
	df["Data"] = pd.to_datetime(df["Data"])
	
	# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
	# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
	df['Lookup'] = df["Data"].dt.strftime('%d.%m.%Y') + df["Interval"].astype(str)
	df.to_excel(file_path, index=False)
	return dataset

def predicting_exporting_Consumption_Solina():
	# Creating the forecast_dataset df
	data = pd.read_csv("./Solina/Solcast/Alba_Iulia_raw.csv")
	forecast_dataset = pd.read_excel("./Solina/Consumption/Input_Consumption_Solina.xlsx")
	# Convert 'period_end' in santimbru to datetime
	data['period_end'] = pd.to_datetime(data['period_end'], errors='coerce')
	# Extract just the date part in the desired format (as strings)
	dates = data['period_end'].dt.strftime('%Y-%m-%d')
	# Write the dates to the Input file
	forecast_dataset['Data'] = dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	forecast_dataset['Data'].fillna(method='bfill', inplace=True)
	# Completing the Interval column
	intervals = data["period_end"].dt.hour
	forecast_dataset["Interval"] = intervals
	# Replace NaNs in the 'Interval' column with 0
	forecast_dataset['Interval'].fillna(0, inplace=True)
	# Completing the Temperatura column
	forecast_dataset["Temperatura"] = data["air_temp"].values

	# Predict on forecast data
	forecast_dataset["Month"] = pd.to_datetime(forecast_dataset.Data).dt.month
	forecast_dataset["WeekDay"] = pd.to_datetime(forecast_dataset.Data).dt.weekday
	forecast_dataset["Holiday"] = 0
	for holiday in forecast_dataset["Data"].unique():
			if holiday in holidays.ds.values:
					forecast_dataset["Holiday"][forecast_dataset["Data"] == holiday] = 1
	dataset = forecast_dataset.copy()
	# Restructuring the dataset
	forecast_dataset = forecast_dataset[["WeekDay", "Month", "Holiday", "Interval", "Temperatura"]]
	# Loading the model
	xgb_loaded = joblib.load("./Solina/Consumption/XGB_Consumption_Temperature.pkl")
	preds = xgb_loaded.predict(forecast_dataset.values)
	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./Solina/Consumption/Results/Results_Consumption_Solina.xlsx")
	worksheet = workbook.add_worksheet("Prediction_Consumption")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in preds:
			worksheet.write(row, col + 2, round(value,3))
			row +=1
	row = 1
	for Data, Interval in zip(dataset.Data, dataset.Interval):
			worksheet.write(row, col + 0, Data, date_format)
			worksheet.write(row, col + 1, Interval)
			row +=1

	workbook.close()
	# Formatting the Results file
	# Step 1: Open the Excel file
	file_path = "./Solina/Consumption/Results/Results_Consumption_Solina.xlsx"
	workbook = load_workbook(filename=file_path)
	worksheet = workbook.active  # Adjust the sheet name as necessary

	# Step 2: Directly round the values in column C and write them back
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value is not None:  # Check if the cell is not empty
			# Round the value to 3 decimal places and write it back to column C
			worksheet.cell(row, 3).value = round(original_value, 3)
		
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value < 0.01:  # Check if the value is less than 0.01
			# Residual values are rounded to 0.000
			worksheet.cell(row, 3).value = 0
	# Save the workbook with the rounded values
	workbook.save(filename=file_path)
	workbook.close()
	return forecast_dataset

def predicting_exporting_RAAL():
	# Creating the forecast_dataset df
	data = pd.read_csv("./RAAL/Solcast/Prundu_raw.csv")
	forecast_dataset = pd.read_excel("./RAAL/Production/Input_RAAL.xlsx", sheet_name="Forecast_Dataset")
	# Convert 'period_end' in santimbru to datetime
	data['period_end'] = pd.to_datetime(data['period_end'], errors='coerce')
	# Extract just the date part in the desired format (as strings)
	dates = data['period_end'].dt.strftime('%Y-%m-%d')
	# Write the dates to the Input file
	forecast_dataset['Data'] = dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	forecast_dataset['Data'].fillna(method='bfill', inplace=True)
	# Completing the Interval column
	intervals = data["period_end"].dt.hour
	forecast_dataset["Interval"] = intervals
	# Replace NaNs in the 'Interval' column with 0
	forecast_dataset['Interval'].fillna(0, inplace=True)
	# Completing the Temperatura column
	forecast_dataset["Temperatura"] = data["air_temp"].values
	# Completing the GHI column
	forecast_dataset["Radiatie"] = data["ghi"].values
	# Completing the Nori column
	forecast_dataset["Nori"] = data["cloud_opacity"].values

	forecast_dataset["Month"] = pd.to_datetime(forecast_dataset.Data).dt.month
	# forecast_dataset.drop("Nori", axis=1, inplace=True)
	dataset = forecast_dataset.copy()
	forecast_dataset = forecast_dataset.drop("Data", axis=1)

	# Loading the model
	xgb_loaded = joblib.load("./RAAL/Production/rs_xgb_RAAL_prod.pkl")
	preds = xgb_loaded.predict(forecast_dataset.values)

	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./RAAL/Production/Results_Production_xgb_RAAL.xlsx")
	worksheet = workbook.add_worksheet("Production_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")

	for value in preds:
			worksheet.write(row, col + 2, round(value,3))
			row +=1
	row = 1
	for Data, Interval in zip(dataset.Data, dataset.Interval):
			worksheet.write(row, col + 0, Data, date_format)
			worksheet.write(row, col + 1, Interval)
			row +=1

	workbook.close()
	# Formatting the Results file
	# Step 1: Open the Excel file
	file_path = "./RAAL/Production/Results_Production_xgb_RAAL.xlsx"
	workbook = load_workbook(filename=file_path)
	worksheet = workbook['Production_Predictions']  # Adjust the sheet name as necessary

	# Step 2: Directly round the values in column C and write them back
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value is not None:  # Check if the cell is not empty
			# Round the value to 3 decimal places and write it back to column C
			worksheet.cell(row, 3).value = round(original_value, 3)
		
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value < 0.01:  # Check if the value is less than 0.01
			# Residual values are rounded to 0.000
			worksheet.cell(row, 3).value = 0
	# Save the workbook with the rounded values
	workbook.save(filename=file_path)
	workbook.close()
	return dataset

def predicting_exporting_Consumption_RAAL():
	# Creating the forecast_dataset df
	data = pd.read_csv("./RAAL/Solcast/Prundu_raw.csv")
	forecast_dataset = pd.read_excel("./RAAL/Production/Input_RAAL.xlsx", sheet_name="Forecast_Dataset")
	# Convert 'period_end' in santimbru to datetime
	data['period_end'] = pd.to_datetime(data['period_end'], errors='coerce')
	# Extract just the date part in the desired format (as strings)
	dates = data['period_end'].dt.strftime('%Y-%m-%d')
	# Write the dates to the Input file
	forecast_dataset['Data'] = dates.values
	# Fill NaNs in the 'Data' column with next valid observation
	forecast_dataset['Data'].fillna(method='bfill', inplace=True)
	# Completing the Interval column
	intervals = data["period_end"].dt.hour
	forecast_dataset["Interval"] = intervals
	# Replace NaNs in the 'Interval' column with 0
	forecast_dataset['Interval'].fillna(0, inplace=True)
	# Completing the Temperatura column
	forecast_dataset["Temperatura"] = data["air_temp"].values
	
	# Predict on forecast data
	dataset = forecast_dataset.copy()
	
	forecast_dataset["Month"] = pd.to_datetime(forecast_dataset.Data).dt.month
	forecast_dataset["WeekDay"] = pd.to_datetime(forecast_dataset.Data).dt.weekday
	forecast_dataset["Holiday"] = 0
	for holiday in forecast_dataset["Data"].unique():
		if holiday in holidays.ds.values:
			forecast_dataset["Holiday"][forecast_dataset["Data"] == holiday] = 1

	# Restructuring the dataset
	forecast_dataset = forecast_dataset[["WeekDay", "Month", "Holiday", "Interval", "Temperatura"]]
	# Loading the model
	xgb_loaded = joblib.load("./RAAL/Consumption/XGB_Consumption_RAAL.pkl")
	preds = xgb_loaded.predict(forecast_dataset.values)
	#Exporting Results to Excel
	workbook = xlsxwriter.Workbook("./RAAL/Consumption//Results/Results_Consumption_RAAL.xlsx")
	worksheet = workbook.add_worksheet("Consumption_Predictions")
	date_format = workbook.add_format({'num_format':'dd.mm.yyyy'})
	# Define a format for cells with three decimal places
	decimal_format = workbook.add_format({'num_format': '0.000'})
	row = 1
	col = 0
	worksheet.write(0,0,"Data")
	worksheet.write(0,1,"Interval")
	worksheet.write(0,2,"Prediction")
	for value in preds:
		worksheet.write(row, col + 2, round(value,3))
		row +=1
	row = 1
	for Data, Interval in zip(dataset.Data, dataset.Interval):
		worksheet.write(row, col + 0, Data, date_format)
		worksheet.write(row, col + 1, Interval)
		row +=1

	workbook.close()
	# Formatting the Results file
	# Step 1: Open the Excel file
	file_path = "./RAAL/Consumption//Results/Results_Consumption_RAAL.xlsx"
	workbook = load_workbook(filename=file_path)
	worksheet = workbook.active  # Adjust the sheet name as necessary

	# Step 2: Directly round the values in column C and write them back
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value is not None:  # Check if the cell is not empty
			# Round the value to 3 decimal places and write it back to column C
			worksheet.cell(row, 3).value = round(original_value, 3)
		
	for row in range(2, worksheet.max_row + 1):
		original_value = worksheet.cell(row, 3).value  # Column C is the 3rd column
		if original_value < 0.01:  # Check if the value is less than 0.01
			# Residual values are rounded to 0.000
			worksheet.cell(row, 3).value = 0
	# Save the workbook with the rounded values
	workbook.save(filename=file_path)
	workbook.close()
	return dataset

def render_consumption_forecast():
	st.write("Consumption Forecast Section")

	# Allow the user to choose between Consumption and Production
	consumption_place = st.radio("Choose Consumption:", options=["Solina", "RAAL"], index=None)
	if consumption_place == "Solina":
		if st.button("Submit"):
			fetching_Solina_data()
			df = predicting_exporting_Consumption_Solina()
			st.dataframe(df)
			file_path = './Solina/Consumption/Results/Results_Consumption_Solina.xlsx'
			with open(file_path, "rb") as f:
				excel_data = f.read()

			# Create a download link
			b64 = base64.b64encode(excel_data).decode()
			button_html = f"""
				 <a download="Consumption_Forecast_Solina.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
				 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
				 </a> 
				 """
			st.markdown(button_html, unsafe_allow_html=True)
	else:
		if st.button("Submit"):
			fetching_RAAL_data()
			df=predicting_exporting_Consumption_RAAL()
			st.dataframe(df)
			file_path = './RAAL/Consumption/Results/Results_Consumption_RAAL.xlsx'
			with open(file_path, "rb") as f:
				excel_data = f.read()

			# Create a download link
			b64 = base64.b64encode(excel_data).decode()
			button_html = f"""
				 <a download="Consumption_Forecast_RAAL.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
				 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
				 </a> 
				 """
			st.markdown(button_html, unsafe_allow_html=True)

def render_production_forecast():
	st.write("Production Forecast Section")

	# Allow the user to choose between Consumption and Production
	PVPP = st.radio("Choose PVPP:", options=["Solina", "RAAL", "Astro", "Imperial"], index=None)

	if PVPP == "Solina":
		# Submit button
		if st.button('Submit'):
			# Fetching the data from Solcast
			fetching_Solina_data()
			# Your code to generate the forecast
			df = predicting_exporting_Solina()
			st.dataframe(df)
			st.success('Forecast Ready', icon="✅")
			file_path = './Solina/Production/Results_Production_xgb.xlsx'
			with open(file_path, "rb") as f:
				excel_data = f.read()

				# Create a download link
				b64 = base64.b64encode(excel_data).decode()
				button_html = f"""
					 <a download="Production_Forecast_Solina.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
					 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
					 </a> 
					 """
				st.markdown(button_html, unsafe_allow_html=True)
	elif PVPP == "RAAL":
		# Submit button
		if st.button("Submit"):
			# Fetching the Solcast data
			fetching_RAAL_data()
			df = predicting_exporting_RAAL()
			st.dataframe(df)
			st.success('Forecast Ready', icon="✅")
			file_path = './RAAL/Production/Results_Production_xgb_RAAL.xlsx'
			with open(file_path, "rb") as f:
				excel_data = f.read()

				# Create a download link
				b64 = base64.b64encode(excel_data).decode()
				button_html = f"""
					 <a download="Production_Forecast.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
					 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
					 </a> 
					 """
				st.markdown(button_html, unsafe_allow_html=True)
	elif PVPP == "Astro":
		st.subheader("Default Forecasting", divider = "red")
		# Submit button
		if st.button("Submit"):
			# Fetching the Solcast data
			fetching_Astro_Imperial_data()
			fetching_Astro_Imperial_data_15min()

			df = predicting_exporting_Astro()
			st.dataframe(df)
			st.success('Forecast Ready', icon="✅")
			file_path = './Astro/Results_Production_Astro_xgb.xlsx'
			with open(file_path, "rb") as f:
				excel_data = f.read()

				# Create a download link
				b64 = base64.b64encode(excel_data).decode()
				button_html = f"""
					 <a download="Production_Forecast_Astro.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
					 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
					 </a> 
					 """
				st.markdown(button_html, unsafe_allow_html=True)
			# Updating the Weather Input 30min granularity
			file_path_input = './Astro/Solcast/Bontida_raw_30min.csv'
			data = pd.read_csv(file_path_input)
			forecast_dataset = pd.read_excel("./Astro/Input_Astro_30min.xlsx")
			# Convert 'period_end' to datetime in UTC
			data['period_end'] = pd.to_datetime(data['period_end'], errors='coerce', utc=True)

			# Manually adjust the time by adding two hours to 'period_end'
			data['period_end'] = data['period_end'] + pd.DateOffset(hours=3)

			# Extract just the date part in the desired format (as strings)
			dates = data['period_end'].dt.strftime('%Y-%m-%d')
			# Write the dates to the Input file
			forecast_dataset['Data'] = dates.values

			# Fill NaNs in the 'Data' column with next valid observation
			forecast_dataset['Data'].fillna(method='bfill', inplace=True)

			# Completing the Interval column
			intervals = data["period_end"].dt.hour
			forecast_dataset["Interval"] = intervals

			# Replace NaNs in the 'Interval' column with 0
			forecast_dataset['Interval'].fillna(0, inplace=True)

			# Completing the Temperatura column
			forecast_dataset["Temperatura"] = data["air_temp"].values

			# Completing the GHI column
			forecast_dataset["Radiatie"] = data["ghi"].values

			# Completing the Nori column
			forecast_dataset["Nori"] = data["cloud_opacity"].values

			# Find indices where 'Interval' equals 1
			indices_of_ones = forecast_dataset.index[forecast_dataset['Interval'] == 1].tolist()

			# Check if there are at least two '1's and replace the first '0' after the second '1'
			if len(indices_of_ones) >= 2:
				second_one_index = indices_of_ones[1]  # Get the index of the second '1'
				# Find the next '0' after the second '1'
				for i in range(second_one_index + 1, len(data)):
					if forecast_dataset.at[i, 'Interval'] == 0:
						forecast_dataset.at[i, 'Interval'] = 2
						break  # Stop after replacing the first '0' to avoid affecting further data

			# Ensure the data is sorted
			forecast_dataset.sort_values(by=['Data', 'Interval'], inplace=True)

			# Initialize the 'Half' column
			forecast_dataset['Half'] = 1  # Start by default with 1

			# Iterate through the DataFrame to manually adjust 'Half'
			prev_interval = None
			count = 0
			for index, row in forecast_dataset.iterrows():
				current_interval = row['Interval']
				if current_interval == prev_interval:
					count += 1
				else:
					count = 1

				# Reset count if it exceeds 2
				if count > 2:
					count = 1

				# Assign 'Half' based on count
				forecast_dataset.at[index, 'Half'] = count

				# Update the previous interval
				prev_interval = current_interval
			# Ensure the 'Data' column is in datetime format
			forecast_dataset["Data"] = pd.to_datetime(forecast_dataset["Data"])
			
			# Create the 'Lookup' column by concatenating the 'Data' and 'Interval' columns
			# Format the 'Data' column as a string in 'dd.mm.yyyy' format for concatenation
			forecast_dataset['Lookup'] = forecast_dataset["Data"].dt.strftime('%d.%m.%Y') + forecast_dataset["Interval"].astype(str) + forecast_dataset["Half"].astype(str)
			forecast_dataset.to_excel("./Astro/Input_Astro_30min.xlsx", index=False)

			with open("./Astro/Input_Astro_30min.xlsx", "rb") as f:
				excel_data = f.read()

				# Create a download link
				b64 = base64.b64encode(excel_data).decode()
				button_html = f"""
					 <a download="Input_Forecast_Astro_30min.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
					 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Input Forecast</button>
					 </a> 
					 """
				st.markdown(button_html, unsafe_allow_html=True)

			creating_prediction_dataset_Astro()
			with open("./Astro/predictions_dataset.xlsx", "rb") as f:
				excel_data = f.read()

				# Create a download link
				b64 = base64.b64encode(excel_data).decode()
				button_html = f"""
					 <a download="Predictions_Dataset_Astro.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
					 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Predictions Dataset</button>
					 </a> 
					 """
				st.markdown(button_html, unsafe_allow_html=True)
			st.dataframe(predicting_exporting_Astro_15min())
			with open("./Astro/Results_Production_Astro_xgb_15min.xlsx", "rb") as f:
				excel_data = f.read()

				# Create a download link
				b64 = base64.b64encode(excel_data).decode()
				button_html = f"""
					 <a download="Predictions_Astro_15min.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
					 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Predictions Astro 15min</button>
					 </a> 
					 """
				st.markdown(button_html, unsafe_allow_html=True)

		st.subheader("Forecasting with Real-Time Production:", divider = "red")
		uploaded_file = st.file_uploader("Upload Real-Time Production Data File", type=['csv', 'xlsx'])
		if uploaded_file is not None:
			if uploaded_file.type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':  # Excel file
				real_time_data = pd.read_excel(uploaded_file, skiprows=4, names=["Timestamp", "Power"])
			elif uploaded_file.type == 'text/csv':  # CSV file
				real_time_data = pd.read_csv(uploaded_file, skiprows=4, names=["Timestamp", "Power"])
			# Process the data (e.g., display first few rows)
			if real_time_data is not None:
				st.write("Uploaded file:")
				st.write(real_time_data.head())
				# Do data processing here (e.g., remove timezone, change "--" to 0, convert power to energy, etc.)
				# Replace '--' with 0 in the 'Power' column
				real_time_data['Power'] = real_time_data['Power'].replace('--', 0)

				# Convert the 'Power' column to numeric (will convert invalid parsing to NaN)
				real_time_data['Power'] = pd.to_numeric(real_time_data['Power'], errors='coerce').fillna(0)

				# Assuming the 'Power' data needs to be in megawatts and currently in watts
				real_time_data['Power_MW'] = real_time_data['Power'] / 1000000  # Convert W to MW

				# Ensure the 'Timestamp' column is in datetime format
				real_time_data['Timestamp'] = pd.to_datetime(real_time_data['Timestamp'])

				# Shift the 'Timestamp' column by 3 hours forward
				real_time_data['Timestamp'] = real_time_data['Timestamp'] + pd.Timedelta(hours=3)

				# Calculate the average of current and next power readings
				real_time_data['Next_Power_MW'] = real_time_data['Power_MW'].shift(1)  # Shift upwards to get the next reading in the column

				# Calculate the average power
				real_time_data['Average_Power_MW'] = (real_time_data['Power_MW'] + real_time_data['Next_Power_MW']) / 2

				# Calculate the energy for each interval
				real_time_data['Energy_MWh'] = real_time_data['Average_Power_MW'] * 0.25
				
				real_time_data.to_csv("./Astro/real-time_data_Astro.csv", index = False)
			else:
				st.write("Unsupported file format. Please upload a CSV or Excel file.")

			st.dataframe(real_time_data)

		if st.button("Forecast Real-Time"):
			fetching_Astro_Imperial_data_past_15min()
			predicting_exporting_Astro_Intraday_15min(real_time_data)
			# Downloading the Predictions Results
			file_path = "./Astro/Results_Production_Astro_xgb_intraday_15min.xlsx"
			with open(file_path, "rb") as f:
				excel_data = f.read()

				# Create a download link
				b64 = base64.b64encode(excel_data).decode()
				button_html = f"""
					 <a download="Production_Forecast_Astro_Intraday_15min.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
					 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results Intraday 15min</button>
					 </a> 
					 """
				st.markdown(button_html, unsafe_allow_html=True)

	elif PVPP == "Imperial":
		# Default 15 min Forecasting
		st.subheader("Default Forecasting", divider = "blue")
		# Submit button
		if st.button("Submit"):
			# Fetching the Solcast data
			# fetching_Astro_Imperial_data()
			df = predicting_exporting_Imperial()
			st.dataframe(df)
			st.success('Forecast Ready', icon="✅")
			file_path = './Imperial/Results_Production_Imperial_xgb.xlsx'
			with open(file_path, "rb") as f:
				excel_data = f.read()

				# Create a download link
				b64 = base64.b64encode(excel_data).decode()
				button_html = f"""
					 <a download="Production_Forecast_Imperial.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
					 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results</button>
					 </a> 
					 """
				st.markdown(button_html, unsafe_allow_html=True)
			st.dataframe(predicting_exporting_Imperial_15min())
			file_path = './Imperial/Results_Production_Imperial_xgb_15min.xlsx'
			with open(file_path, "rb") as f:
				excel_data = f.read()

				# Create a download link
				b64 = base64.b64encode(excel_data).decode()
				button_html = f"""
					 <a download="Production_Forecast_Imperial_15min.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
					 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results 15min</button>
					 </a> 
					 """
				st.markdown(button_html, unsafe_allow_html=True)

		# Forecasting using the Real-Time production data		
		st.subheader("Forecasting with Real-Time Production:", divider = "blue")
		uploaded_files = st.file_uploader("Upload Real-Time Production Data Files", type=['csv', 'xlsx'], accept_multiple_files=True)
		if uploaded_files:
			count = 0
			# Process each uploaded file
			combined_data = pd.DataFrame()
			for uploaded_file in uploaded_files:
				count = count + 1
				if uploaded_file.type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet':  # Excel file
					real_time_data = pd.read_excel(uploaded_file, skiprows=4, names=["Timestamp", "Power"])
				elif uploaded_file.type == 'text/csv':  # CSV file
					real_time_data = pd.read_csv(uploaded_file, skiprows=4, names=["Timestamp", "Power"])
				
				# Do data processing here (e.g., remove timezone, change "--" to 0, convert power to energy, etc.)
	 
				# Replace '--' with 0 in the 'Power' column
				real_time_data['Power'] = real_time_data['Power'].replace('--', 0)

				# Convert the 'Power' column to numeric (will convert invalid parsing to NaN)
				real_time_data['Power'] = pd.to_numeric(real_time_data['Power'], errors='coerce').fillna(0)

				# Assuming the 'Power' data needs to be in megawatts and currently in watts
				real_time_data['Power_MW'] = real_time_data['Power'] / 1000000  # Convert W to MW

				# Ensure the 'Timestamp' column is in datetime format
				real_time_data['Timestamp'] = pd.to_datetime(real_time_data['Timestamp'])

				# Shift the 'Timestamp' column by 3 hours forward
				real_time_data['Timestamp'] = real_time_data['Timestamp'] + pd.Timedelta(hours=3)

				# Calculate the average of current and next power readings
				real_time_data['Next_Power_MW'] = real_time_data['Power_MW'].shift(1)  # Shift upwards to get the next reading in the column

				# Calculate the average power
				real_time_data['Average_Power_MW'] = (real_time_data['Power_MW'] + real_time_data['Next_Power_MW']) / 2

				# Calculate the energy for each interval
				real_time_data['Energy_MWh'] = real_time_data['Average_Power_MW'] * 0.25
				
				real_time_data.to_csv("./Imperial/real-time_data_Imperial{}.csv".format(count), index = False)

		if st.button("Forecast Real-Time"):
			fetching_Astro_Imperial_data_past_15min()
			# Creating the real-time production dataframe for Imperial
			real_time_data = pd.read_csv("./Imperial/real-time_data_Imperial1.csv")
			second_data = pd.read_csv("./Imperial/real-time_data_Imperial2.csv")
			# Iterate over columns to add from the second DataFrame
			for column in ['Power', 'Power_MW', 'Next_Power_MW', 'Average_Power_MW', 'Energy_MWh']:
				# Sum values from the second DataFrame and add them to the respective column in the final DataFrame
				real_time_data[column] += second_data[column]
			real_time_data.to_excel("./Imperial/Real-Time_forecast_dataset.xlsx", index = False)
			preds = predicting_exporting_Imperial_Intraday_15min(real_time_data)
			
			# Downloading the Predictions Results
			file_path = "./Imperial/Results_Production_Imperial_xgb_intraday_15min.xlsx"
			with open(file_path, "rb") as f:
				excel_data = f.read()

				# Create a download link
				b64 = base64.b64encode(excel_data).decode()
				button_html = f"""
					 <a download="Production_Forecast_Imperial_Intraday_15min.xlsx" href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download>
					 <button kind="secondary" data-testid="baseButton-secondary" class="st-emotion-cache-12tniow ef3psqc12">Download Forecast Results Intraday 15min</button>
					 </a> 
					 """
				st.markdown(button_html, unsafe_allow_html=True)

def render_forecast_page():
	
	# Web App Title
	st.markdown('''
	# **The Forecast Section**

	''')

	# Allow the user to choose between Consumption and Production
	forecast_type = st.radio("Choose Forecast Type:", options=["Consumption", "Production", "Transavia"])

	if forecast_type == "Consumption":
		cleaning_input_files()
		render_consumption_forecast()
	elif forecast_type == "Production":
		render_production_forecast()
	elif forecast_type == "Transavia":
		print("Transavia page")
		render_Transavia_page()
		

#======================================================BALANGING MARKET===================================================================================================

def render_balancing_market_page():
	
	# Web App Title
	st.markdown('''
	# **The Balancing Market Section**

	''')
	st.divider()